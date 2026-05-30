from __future__ import annotations

import copy
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook


ACTION_COLUMNS = [
    "action_id",
    "action_type",
    "object_type",
    "portfolio_id",
    "portfolio_name",
    "entity_id",
    "campaign_id",
    "ad_group_id",
    "ad_id",
    "keyword_id",
    "product_target_id",
    "campaign_name",
    "ad_group_name",
    "sku",
    "keyword",
    "match_type",
    "bidding_strategy",
    "budget",
    "bid",
    "placement",
    "percentage",
    "negative_type",
    "state",
    "block_id",
    "block_index",
    "block_title",
    "reason",
    "source_text",
    "source_row_index",
]
SKIPPED_COLUMNS = ACTION_COLUMNS + ["skip_reason"]
VALIDATION_COLUMNS = ["check", "status", "detail", "action_id", "generated_at"]

BULK_HEADER_ALIASES = {
    "product": ["Product", "产品"],
    "entity": ["Entity", "Entity Level", "Record Type", "实体层级", "实体"],
    "entity_id": ["Entity ID", "Entity Id", "Record ID", "Record Id", "实体编号"],
    "operation": ["Operation", "操作"],
    "portfolio_id": ["Portfolio ID", "Portfolio Id", "Portfolio Identifier", "广告组合编号", "广告组合 ID", "广告组合ID"],
    "portfolio_name": ["Portfolio Name", "Portfolio Name (Informational only)", "广告组合名称", "广告组合名称（仅供参考）"],
    "campaign_id": ["Campaign ID", "Campaign Id", "Campaign Identifier", "广告活动编号", "广告活动 ID", "广告活动ID"],
    "campaign_name": ["Campaign Name", "广告活动名称"],
    "ad_group_id": ["Ad Group ID", "Ad Group Id", "Ad Group Identifier", "广告组编号", "广告组 ID", "广告组ID"],
    "ad_group_name": ["Ad Group Name", "广告组名称"],
    "ad_id": ["Ad ID", "Ad Id", "Ad Identifier", "广告编号", "广告 ID", "广告ID"],
    "keyword_id": ["Keyword ID", "Keyword Id", "Keyword Identifier", "关键词编号", "关键词 ID", "关键词ID"],
    "product_target_id": ["Product Targeting ID", "Product Targeting Id", "Product Targeting Identifier", "Targeting ID", "Targeting Id", "商品投放 ID", "商品投放ID", "投放 ID", "投放ID", "拓展商品投放编号", "拓展商品投放 ID", "拓展商品投放ID", "扩展商品投放编号", "扩展商品投放 ID", "扩展商品投放ID"],
    "targeting_type": ["Targeting Type", "投放类型"],
    "state": ["State", "Status", "状态"],
    "bidding_strategy": ["Bidding Strategy", "Bidding Strategy Type", "竞价方案", "竞价策略", "竞价方式"],
    "daily_budget": ["Daily Budget", "Budget", "每日预算", "预算"],
    "sku": ["SKU"],
    "ad_group_default_bid": ["Ad Group Default Bid", "广告组默认竞价"],
    "bid": ["Bid", "竞价"],
    "keyword_text": ["Keyword Text", "Keyword", "Targeting Text", "Targeting", "关键词文本", "关键词", "关键词或商品投放", "关键词或商品定位", "投放内容", "投放"],
    "match_type": ["Match Type", "匹配类型"],
    "product_targeting_expression": ["Product Targeting Expression", "Targeting Expression", "Expression", "Resolved Expression", "商品投放表达式", "关键词投放表达式", "投放表达式", "定位表达式", "拓展商品投放名称", "拓展商品投放名称（仅供参考）", "扩展商品投放名称", "扩展商品投放名称（仅供参考）"],
    "placement": ["Placement", "Placement Type", "广告位", "广告位置", "位置"],
    "percentage": ["Percentage", "Percentage Increase", "Increase by percentage", "Placement Percentage", "百分比", "广告位百分比", "加价百分比"],
}

BASE_REQUIRED_BULK_FIELDS = ["product", "entity", "operation", "campaign_name", "state"]
CONTROLLED_BULK_FIELDS = ["product", "entity", "operation", "targeting_type", "state", "match_type", "bidding_strategy"]
SUPPORTED_ACTIONS = {
    "create_campaign",
    "create_ad_group",
    "create_product_ad",
    "create_keyword_target",
    "create_product_target",
    "update_budget",
    "update_bid",
    "update_product_targeting_bid",
    "update_product_target_bid",
    "add_negative_exact",
    "add_negative_phrase",
    "update_placement",
    "pause_campaign",
    "pause_target",
    "create_keyword_campaign",
    "create_product_targeting_campaign",
    "update_keyword_bid",
    "update_product_targeting_bid",
    "update_campaign_budget",
}
MATCH_TYPES = {"exact", "phrase", "broad"}
PRODUCT_TARGETING_BID_ACTIONS = {"update_product_targeting_bid", "update_product_target_bid"}
UPDATE_COPY_ACTIONS = {"update_bid", "update_product_targeting_bid", "update_product_target_bid", "update_budget", "pause_campaign", "pause_target"}
BULK_UPDATE_ACTIONS = {"update_bid", "update_budget", "update_placement", "pause_campaign", "pause_target"} | PRODUCT_TARGETING_BID_ACTIONS


@dataclass
class BulkTemplate:
    workbook: object
    sheet: object
    header_row_idx: int
    headers: list[str]
    header_map: dict[str, int]
    data: pd.DataFrame
    language: str
    labels: dict[str, dict[str, str]]
    style_row: list[object]


@dataclass(frozen=True)
class BulkIndexes:
    campaigns: pd.DataFrame
    ad_groups: pd.DataFrame
    targets: pd.DataFrame
    negatives: pd.DataFrame
    portfolios: pd.DataFrame


@dataclass(frozen=True)
class GenerationResult:
    actions: pd.DataFrame
    skipped_actions: pd.DataFrame
    validation: pd.DataFrame
    bulk_upload: bytes | None
    summary: dict[str, object]


def run_generation(*, bulk_template: BinaryIO | bytes | None, requirement_text: str) -> GenerationResult:
    validation_rows: list[dict[str, object]] = []
    template: BulkTemplate | None = None
    indexes = _empty_indexes()
    actions = _empty_actions()
    skipped = _empty_skipped()
    executable = _empty_actions()

    if bulk_template is None:
        validation_rows.append(_validation_row("文件读取", "error", "必须上传 Sponsored Products Bulk xlsx 文件"))
    else:
        try:
            template = read_bulk_template(bulk_template)
            indexes = build_indexes(template)
            validation_rows.extend(_validate_template(template))
            validation_rows.append(_validation_row("文件读取", "ok", f"已识别 Sheet：{template.sheet.title}"))
        except Exception as exc:  # noqa: BLE001
            validation_rows.append(_validation_row("文件读取", "error", f"Bulk xlsx 读取失败：{exc}"))

    requirement_text = requirement_text or ""
    operation_blocks = split_operation_blocks(requirement_text)
    if not requirement_text.strip():
        validation_rows.append(_validation_row("需求解析", "error", "未输入需求"))
    else:
        actions = parse_requirement(requirement_text)
        validation_rows.append(_validation_row("需求读取", "ok", f"已读取需求 {len(requirement_text)} 字"))
        parse_failures = _parse_failure_prechecks(operation_blocks, actions)
        for failure in parse_failures:
            status = "warning" if not actions.empty else "error"
            validation_rows.append(
                _validation_row(
                    f"操作{failure['block_index']}解析",
                    status,
                    f"{failure['missing_fields']}；{failure['suggestion']}",
                    str(failure.get("block_id", "")),
                )
            )
        if actions.empty:
            preview = requirement_text[:500].replace("\n", "\\n")
            validation_rows.append(_validation_row("需求解析", "error", f"未识别到动作。原始需求前500字：{preview}"))
        else:
            validation_rows.append(_validation_row("需求解析", "ok", f"已解析 {len(actions)} 行标准操作单 JSON"))
            executable, skipped, action_validation = validate_actions(actions, template, indexes)
            validation_rows.extend(action_validation)

    bulk_upload = None
    failure_reasons = _failure_reasons(validation_rows)
    if executable.empty:
        if skipped.empty:
            detail = failure_reasons[0] if failure_reasons else "actions.csv 为空，无法生成 Bulk 上传数据"
            validation_rows.append(_validation_row("Bulk 输出", "error", f"未生成 bulk_upload.xlsx：{detail}"))
        else:
            reasons = _joined_unique(skipped["skip_reason"].tolist())
            validation_rows.append(_validation_row("Bulk 输出", "error", f"未生成 bulk_upload.xlsx：{reasons}"))
    elif template is None:
        validation_rows.append(_validation_row("Bulk 输出", "error", "未生成 bulk_upload.xlsx：缺少可用 Bulk 模板"))
    else:
        bulk_upload = build_bulk_upload(template, executable)
        if failure_reasons:
            validation_rows.append(_validation_row("Bulk 输出", "warning", f"部分操作块未生成：{'; '.join(failure_reasons)}"))
        validation_rows.append(_validation_row("Bulk 输出", "ok", f"已生成 {len(executable)} 行 Bulk 上传数据"))

    validation = pd.DataFrame(validation_rows, columns=VALIDATION_COLUMNS)
    failure_reasons = _failure_reasons(validation_rows)
    prechecks = _precheck_summary(actions, executable, skipped, operation_blocks)
    block_counts = _precheck_counts(prechecks)
    return GenerationResult(
        actions=actions,
        skipped_actions=skipped,
        validation=validation,
        bulk_upload=bulk_upload,
        summary={
            "actions": len(actions),
            "skipped": len(skipped),
            "bulk_ready": bulk_upload is not None,
            "validation_passed": not failure_reasons,
            "failure_reasons": failure_reasons,
            "prechecks": prechecks,
            "operation_json": actions.to_dict("records"),
            "total_blocks": block_counts["total"],
            "success_blocks": block_counts["success"],
            "partial_blocks": block_counts["partial"],
            "partial_success": block_counts["partial"],
            "failed_blocks": block_counts["failed"],
            "generated_rows": len(executable),
        },
    )


def read_bulk_template(file: BinaryIO | bytes | str | Path) -> BulkTemplate:
    data = _read_bytes(file)
    workbook = load_workbook(BytesIO(data))
    sheet, header_row_idx, headers = _find_bulk_sheet(workbook)
    if sheet is None or header_row_idx is None:
        raise ValueError("未识别到 Sponsored Products Bulk 表头")
    header_map = _build_header_map(headers)
    dataframe = _sheet_to_dataframe(sheet, header_row_idx, headers)
    labels = _extract_labels(workbook, dataframe, header_map)
    style_row = _capture_style_row(sheet, header_row_idx)
    language = "zh" if any(_has_cjk(header) for header in headers) else "en"
    return BulkTemplate(workbook, sheet, header_row_idx, headers, header_map, dataframe, language, labels, style_row)


def build_indexes(template: BulkTemplate) -> BulkIndexes:
    rows = []
    for source_row_index, row in template.data.iterrows():
        entity = _canonical_entity(_row_value(row, template.header_map, "entity"))
        target_value = _target_value(row, template.header_map)
        product_target_id = _row_value(row, template.header_map, "product_target_id")
        if entity == "keyword_target" and (product_target_id or _looks_like_asin(target_value)):
            entity = "product_target"
        rows.append(
            {
                "entity": entity,
                "entity_id": _entity_id_for_row(row, template.header_map),
                "portfolio_id": _row_value(row, template.header_map, "portfolio_id"),
                "portfolio_name": _row_value(row, template.header_map, "portfolio_name"),
                "campaign_id": _row_value(row, template.header_map, "campaign_id"),
                "campaign_name": _row_value(row, template.header_map, "campaign_name"),
                "ad_group_id": _row_value(row, template.header_map, "ad_group_id"),
                "ad_group_name": _row_value(row, template.header_map, "ad_group_name"),
                "ad_id": _row_value(row, template.header_map, "ad_id"),
                "keyword_id": _row_value(row, template.header_map, "keyword_id"),
                "product_target_id": product_target_id,
                "keyword": target_value,
                "match_type": _canonical_match_type(_row_value(row, template.header_map, "match_type")),
                "negative_type": _canonical_negative_type(_row_value(row, template.header_map, "match_type")),
                "source_row_index": source_row_index,
            }
        )
    index_columns = ["entity", "entity_id", "portfolio_id", "portfolio_name", "campaign_id", "campaign_name", "ad_group_id", "ad_group_name", "ad_id", "keyword_id", "product_target_id", "keyword", "match_type", "negative_type", "source_row_index"]
    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=index_columns)
    df = _fill_parent_names(df)
    campaigns = df[df["campaign_name"].astype(str).str.strip() != ""][["entity", "campaign_name", "entity_id", "campaign_id", "portfolio_id", "portfolio_name", "source_row_index"]].drop_duplicates()
    ad_groups = df[df["ad_group_name"].astype(str).str.strip() != ""][["campaign_name", "campaign_id", "ad_group_name", "entity_id", "ad_group_id"]].drop_duplicates()
    target_mask = df["entity"].isin(["keyword_target", "product_target"])
    target_mask |= (
        df["keyword"].astype(str).str.strip().ne("")
        & df["ad_group_id"].astype(str).str.strip().ne("")
        & df["negative_type"].astype(str).str.strip().eq("")
    )
    targets = df[target_mask][["entity", "campaign_name", "campaign_id", "ad_group_name", "ad_group_id", "entity_id", "keyword_id", "product_target_id", "keyword", "match_type", "source_row_index"]].drop_duplicates()
    negatives = df[df["entity"].isin(["negative_keyword", "campaign_negative_keyword"])][["campaign_name", "campaign_id", "ad_group_name", "ad_group_id", "entity_id", "keyword_id", "keyword", "negative_type"]].drop_duplicates()
    portfolios = _build_portfolio_index(template, df)
    return BulkIndexes(campaigns, ad_groups, targets, negatives, portfolios)


def _build_portfolio_index(template: BulkTemplate, bulk_df: pd.DataFrame) -> pd.DataFrame:
    records: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()

    def add_record(portfolio_name: object, portfolio_id: object) -> None:
        name = _clean_text(portfolio_name)
        portfolio_id_text = _clean_text(portfolio_id)
        if not name and not portfolio_id_text:
            return
        key = (_norm_key(name), _norm_key(portfolio_id_text))
        if key in seen:
            return
        seen.add(key)
        records.append({"portfolio_name": name, "portfolio_id": portfolio_id_text})

    if not bulk_df.empty:
        for _, row in bulk_df.iterrows():
            add_record(row.get("portfolio_name", ""), row.get("portfolio_id", ""))

    for sheet in template.workbook.worksheets:
        header_row_idx, headers, _ = _find_header_row(sheet)
        if header_row_idx is None:
            continue
        header_map = _build_header_map(headers)
        if "portfolio_id" not in header_map and "portfolio_name" not in header_map:
            continue
        for values in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            row_values = list(values[: len(headers)])
            row_values.extend([None] * (len(headers) - len(row_values)))
            if not any(value not in ("", None) for value in row_values):
                continue
            row = pd.Series(row_values[: len(headers)])
            add_record(_row_value(row, header_map, "portfolio_name"), _row_value(row, header_map, "portfolio_id"))

    columns = ["portfolio_name", "portfolio_id"]
    if not records:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(records, columns=columns).drop_duplicates()


def _fill_parent_names(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    filled = df.copy()
    campaign_rows = filled[
        filled["campaign_id"].astype(str).str.strip().ne("")
        & filled["campaign_name"].astype(str).str.strip().ne("")
    ]
    campaign_names = {
        str(key).strip(): str(value).strip()
        for key, value in (
        campaign_rows.drop_duplicates("campaign_id")
        .set_index("campaign_id")["campaign_name"]
        .astype(str)
        .to_dict()
        ).items()
        if str(key).strip() and str(value).strip()
    }
    campaign_ids = {
        str(key).strip(): str(value).strip()
        for key, value in (
        campaign_rows.drop_duplicates("campaign_name")
        .set_index("campaign_name")["campaign_id"]
        .astype(str)
        .to_dict()
        ).items()
        if str(key).strip() and str(value).strip()
    }
    ad_group_rows = filled[
        filled["ad_group_id"].astype(str).str.strip().ne("")
        & filled["ad_group_name"].astype(str).str.strip().ne("")
    ]
    ad_group_names = {
        str(key).strip(): str(value).strip()
        for key, value in (
        ad_group_rows.drop_duplicates("ad_group_id")
        .set_index("ad_group_id")["ad_group_name"]
        .astype(str)
        .to_dict()
        ).items()
        if str(key).strip() and str(value).strip()
    }
    ad_group_campaign_ids = {
        str(key).strip(): str(value).strip()
        for key, value in (
        ad_group_rows.drop_duplicates("ad_group_id")
        .set_index("ad_group_id")["campaign_id"]
        .astype(str)
        .to_dict()
        ).items()
        if str(key).strip() and str(value).strip()
    }

    for idx, row in filled.iterrows():
        campaign_id = str(row.get("campaign_id", "") or "").strip()
        campaign_name = str(row.get("campaign_name", "") or "").strip()
        ad_group_id = str(row.get("ad_group_id", "") or "").strip()
        ad_group_name = str(row.get("ad_group_name", "") or "").strip()
        if not campaign_name and campaign_id in campaign_names:
            filled.at[idx, "campaign_name"] = campaign_names[campaign_id]
        if not campaign_id and campaign_name in campaign_ids:
            filled.at[idx, "campaign_id"] = campaign_ids[campaign_name]
        if not ad_group_name and ad_group_id in ad_group_names:
            filled.at[idx, "ad_group_name"] = ad_group_names[ad_group_id]
        if not campaign_id and ad_group_id in ad_group_campaign_ids:
            filled.at[idx, "campaign_id"] = ad_group_campaign_ids[ad_group_id]
            parent_campaign_name = campaign_names.get(ad_group_campaign_ids[ad_group_id], "")
            if not campaign_name and parent_campaign_name:
                filled.at[idx, "campaign_name"] = parent_campaign_name
    return filled


def split_operation_blocks(requirement_text: str) -> list[dict[str, str]]:
    text = (requirement_text or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return []

    if re.search(r"(?m)^\s*---+\s*$", text):
        parts = re.split(r"(?m)^\s*---+\s*$", text)
    elif re.search(r"(?m)^\s*\d+[\.、)]\s+", text):
        parts = re.split(r"(?m)(?=^\s*\d+[\.、)]\s+)", text)
    elif len(re.findall(r"(?m)^\s*操作\s*[：:]", text)) > 1:
        parts = re.split(r"(?m)(?=^\s*操作\s*[：:])", text)
    else:
        parts = [text]

    blocks: list[dict[str, str]] = []
    for raw_part in parts:
        block_text = _compact_block_text(raw_part)
        if not block_text:
            continue
        block_index = len(blocks) + 1
        blocks.append(
            {
                "block_id": f"B{block_index:03d}",
                "block_index": str(block_index),
                "block_title": _operation_block_title(block_text, block_index),
                "text": block_text,
            }
        )
    return blocks


def parse_requirement(requirement_text: str) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for block in split_operation_blocks(requirement_text):
        block_rows = _parse_requirement_block(block["text"])
        for row in block_rows:
            row["block_id"] = block["block_id"]
            row["block_index"] = block["block_index"]
            row["block_title"] = block["block_title"]
        rows.extend(block_rows)
    return _assign_action_ids(_clean_actions(pd.DataFrame(rows)))


def _parse_requirement_block(requirement_text: str) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for section in _split_sections(requirement_text):
        kind = _section_kind(section["title"], section["body"])
        if kind == "create_keyword_campaign":
            rows.extend(_parse_create_ad_section(section))
        elif kind == "create_product_targeting_campaign":
            rows.extend(_parse_create_product_targeting_campaign_section(section))
        elif kind in {"add_negative_exact", "add_negative_phrase"}:
            rows.extend(_parse_negative_section(section, kind))
        elif kind == "update_campaign_budget":
            row = _parse_budget_section(section)
            if row:
                rows.append(row)
        elif kind == "update_keyword_bid":
            rows.extend(_parse_bid_section(section))
        elif kind in PRODUCT_TARGETING_BID_ACTIONS:
            rows.extend(_parse_product_target_bid_section(section))
        elif kind == "update_placement":
            rows.extend(_parse_placement_section(section))
        elif kind == "pause_campaign":
            row = _parse_pause_campaign_section(section)
            if row:
                rows.append(row)
        elif kind == "pause_target":
            rows.extend(_parse_pause_target_section(section))
    if "预算改为" in requirement_text and not any(row.get("action_type") == "update_budget" and row.get("campaign_name") for row in rows):
        row = _parse_budget_section({"title": "预算改为", "body": requirement_text})
        if row:
            rows.append(row)
    if "竞价改为" in requirement_text and not any(row.get("action_type") == "update_bid" and row.get("campaign_name") and row.get("keyword") for row in rows):
        rows.extend(_parse_bid_section({"title": "竞价改为", "body": requirement_text}))
    return rows


def _compact_block_text(text: str) -> str:
    lines = []
    for raw_line in str(text or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        line = re.sub(r"^\s*\d+[\.、)]\s*", "", line).strip()
        if line:
            lines.append(line)
    return "\n".join(lines).strip()


def _operation_block_title(text: str, block_index: int) -> str:
    for raw_line in str(text or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        operation_match = re.match(r"^操作\s*[：:]\s*(.+)$", line)
        if operation_match:
            return _clean_text(operation_match.group(1)) or f"操作{block_index}"
        return line.rstrip("：:")[:80] or f"操作{block_index}"
    return f"操作{block_index}"


def validate_actions(
    actions: pd.DataFrame,
    template: BulkTemplate | None,
    indexes: BulkIndexes,
) -> tuple[pd.DataFrame, pd.DataFrame, list[dict[str, object]]]:
    executable: list[dict[str, object]] = []
    skipped: list[dict[str, object]] = []
    validation: list[dict[str, object]] = []
    fatal_block_ids: set[str] = set()
    actions = _assign_create_temp_ids(actions)
    create_campaigns = set(actions[actions["action_type"] == "create_campaign"]["campaign_name"].astype(str))
    create_ad_groups = {
        (_norm_key(row["campaign_name"]), _norm_key(row["ad_group_name"]))
        for row in actions[actions["action_type"] == "create_ad_group"].to_dict("records")
    }

    for action in actions.to_dict("records"):
        action = _enrich_action(action, indexes)
        errors = _action_errors(action, template, indexes, create_campaigns, create_ad_groups)
        if errors:
            detail = "；".join(errors)
            skipped.append({**action, "skip_reason": detail})
            if not _is_nonfatal_skip(errors):
                fatal_block_ids.add(str(action.get("block_id", "")))
            validation.append(_validation_row("动作校验", "warning", detail, action.get("action_id", "")))
        else:
            executable.append(action)

    structural_errors = _new_campaign_structure_errors(actions)
    if structural_errors:
        blocked_ids = set()
        for campaign_name, errors in structural_errors.items():
            detail = f"{campaign_name or '未命名广告活动'}：{'；'.join(errors)}"
            validation.append(_validation_row("新建广告结构", "warning", detail))
            for action in actions[actions["campaign_name"] == campaign_name].to_dict("records"):
                if action["action_type"].startswith("create_"):
                    blocked_ids.add(action["action_id"])
                    fatal_block_ids.add(str(action.get("block_id", "")))
                    skipped.append({**action, "skip_reason": "；".join(errors)})
        executable = [action for action in executable if action["action_id"] not in blocked_ids]

    if fatal_block_ids:
        executable = [action for action in executable if str(action.get("block_id", "")) not in fatal_block_ids]

    executable_df = _clean_actions(pd.DataFrame(executable))
    if template is not None and not executable_df.empty:
        executable_df, validation_skipped, block_validation = _validate_executable_blocks(template, executable_df)
        skipped.extend(validation_skipped)
        validation.extend(block_validation)
    skipped_df = _clean_skipped(pd.DataFrame(_dedupe_skipped(skipped)))
    if skipped_df.empty and not executable_df.empty and not _has_errors(validation):
        validation.append(_validation_row("动作校验", "ok", "全部动作校验通过"))
    return executable_df, skipped_df, validation


def _validate_executable_blocks(
    template: BulkTemplate,
    executable: pd.DataFrame,
) -> tuple[pd.DataFrame, list[dict[str, object]], list[dict[str, object]]]:
    valid_blocks: list[pd.DataFrame] = []
    skipped: list[dict[str, object]] = []
    validation: list[dict[str, object]] = []

    for block_id, block_actions in _block_groups(executable):
        block_validation: list[dict[str, object]] = []
        block_validation.extend(_scope_block_validation_rows(block_actions, _required_column_validation(template, block_actions), block_id))
        block_validation.extend(_scope_block_validation_rows(block_actions, _bulk_id_validation(template, block_actions), block_id))
        block_validation.extend(_scope_block_validation_rows(block_actions, _generated_bulk_value_validation(template, block_actions), block_id))
        block_errors = _failure_reasons(block_validation)
        validation.extend(block_validation)
        if block_errors:
            reason = "；".join(block_errors)
            skipped.extend({**action, "skip_reason": reason} for action in block_actions.to_dict("records"))
        else:
            valid_blocks.append(block_actions)

    if valid_blocks:
        return _clean_actions(pd.concat(valid_blocks, ignore_index=True)), skipped, validation
    return _empty_actions(), skipped, validation


def _scope_block_validation_rows(
    block_actions: pd.DataFrame,
    rows: list[dict[str, object]],
    block_id: str,
) -> list[dict[str, object]]:
    block_index = _first_value(block_actions, "block_index") or block_id
    scoped = []
    for row in rows:
        scoped_row = dict(row)
        detail = str(scoped_row.get("detail", ""))
        if scoped_row.get("status") == "error" and not detail.startswith("操作"):
            scoped_row["detail"] = f"操作{block_index}失败：{detail}"
        if not str(scoped_row.get("action_id", "")).strip():
            scoped_row["action_id"] = block_id
        scoped.append(scoped_row)
    return scoped


def build_bulk_upload(template: BulkTemplate, actions: pd.DataFrame) -> bytes:
    sheet = template.sheet
    _clear_data_rows(sheet, template.header_row_idx)
    for action in actions.to_dict("records"):
        if str(action.get("action_type", "")) in UPDATE_COPY_ACTIONS and str(action.get("source_row_index", "")).strip() != "":
            row_values = _source_row_values(template, action)
            values = _controlled_update_values(action, template)
            values.update(_copy_update_values(action, template))
        else:
            values = _bulk_values(action, template)
            row_values = [None for _ in template.headers]
        for field, value in values.items():
            _set_row_value(row_values, template.header_map, field, value)
        sheet.append(row_values)
        _apply_style_to_row(sheet, sheet.max_row, template.style_row)
    output = BytesIO()
    template.workbook.save(output)
    return output.getvalue()


def dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def _precheck_summary(
    actions: pd.DataFrame,
    executable: pd.DataFrame,
    skipped: pd.DataFrame,
    operation_blocks: list[dict[str, str]] | None = None,
) -> list[dict[str, object]]:
    summaries: list[dict[str, object]] = []
    if not actions.empty:
        for block_id, group in _block_groups(actions):
            executable_group = _matching_block_rows(executable, block_id)
            skipped_group = _matching_block_rows(skipped, block_id)
            operation_type = _block_operation_type(group)
            status = _block_status(executable_group, skipped_group)
            missing_fields = _joined_unique(_missing_fields(skipped_group)) if not skipped_group.empty else ""
            unmatched_objects = _joined_unique(_object_labels(skipped_group)) if not skipped_group.empty else ""
            block_index = _first_value(group, "block_index")
            block_title = _first_value(group, "block_title") or _operation_label(operation_type)
            matched_asins = _joined_unique(_asin_labels(executable_group))
            unmatched_asins = _joined_unique(_asin_labels(skipped_group))
            keywords = _joined_unique(_keyword_labels(group))
            asins = _joined_unique(_asin_labels(group))
            suggestion = _suggestion_for_reason(missing_fields or unmatched_objects)
            summaries.append(
                {
                    "block_id": block_id,
                    "block_index": int(block_index) if str(block_index).isdigit() else len(summaries) + 1,
                    "block_title": block_title,
                    "operation_type": _operation_label(operation_type),
                    "action_type": operation_type,
                    "status": status,
                    "campaign_name": _joined_unique(group["campaign_name"].tolist()),
                    "ad_group_name": _joined_unique(group["ad_group_name"].tolist()),
                    "budget": _joined_unique(group["budget"].tolist()),
                    "bid": _joined_unique(group["bid"].tolist()),
                    "placement": _joined_unique(group["placement"].tolist()) if "placement" in group.columns else "",
                    "percentage": _joined_unique(group["percentage"].tolist()) if "percentage" in group.columns else "",
                    "keywords": keywords,
                    "asins": asins,
                    "keyword_count": _count_keyword_rows(group),
                    "asin_count": _count_asin_rows(group),
                    "negative_count": _count_negative_rows(group),
                    "matched_objects": _joined_unique(_object_labels(executable_group)) if not executable_group.empty else "",
                    "unmatched_objects": unmatched_objects,
                    "matched_asins": matched_asins,
                    "unmatched_asins": unmatched_asins,
                    "missing_fields": missing_fields,
                    "suggestion": suggestion,
                    "generated_rows": int(len(executable_group)),
                    "recognized_fields": _recognized_fields(group),
                }
            )

    summaries.extend(_parse_failure_prechecks(operation_blocks or [], actions))
    return sorted(summaries, key=lambda item: int(item.get("block_index", 0) or 0))


def _legacy_precheck_summary(actions: pd.DataFrame, executable: pd.DataFrame, skipped: pd.DataFrame) -> list[dict[str, object]]:
    if actions.empty:
        return []
    summaries = []
    for operation_type, group in _operation_groups(actions):
        executable_group = _matching_operation_rows(executable, operation_type)
        skipped_group = _matching_operation_rows(skipped, operation_type)
        summaries.append(
            {
                "operation_type": _operation_label(operation_type),
                "action_type": operation_type,
                "campaign_name": _joined_unique(group["campaign_name"].tolist()),
                "ad_group_name": _joined_unique(group["ad_group_name"].tolist()),
                "budget": _joined_unique(group["budget"].tolist()),
                "bid": _joined_unique(group["bid"].tolist()),
                "placement": _joined_unique(group["placement"].tolist()) if "placement" in group.columns else "",
                "percentage": _joined_unique(group["percentage"].tolist()) if "percentage" in group.columns else "",
                "keyword_count": _count_keyword_rows(group),
                "asin_count": _count_asin_rows(group),
                "negative_count": _count_negative_rows(group),
                "matched_objects": _joined_unique(_object_labels(executable_group)) if not executable_group.empty else "",
                "unmatched_objects": _joined_unique(_object_labels(skipped_group)) if not skipped_group.empty else "",
                "missing_fields": _joined_unique(_missing_fields(skipped_group)) if not skipped_group.empty else "",
                "generated_rows": int(len(executable_group)),
                "recognized_fields": _recognized_fields(group),
            }
        )
    return summaries


def _block_groups(actions: pd.DataFrame) -> list[tuple[str, pd.DataFrame]]:
    cleaned = _clean_actions(actions)
    if cleaned.empty:
        return []
    if "block_id" not in cleaned.columns or not cleaned["block_id"].astype(str).str.strip().ne("").any():
        return [("B001", cleaned)]
    grouped: list[tuple[str, pd.DataFrame]] = []
    for block_id in cleaned["block_id"].drop_duplicates().tolist():
        block_key = str(block_id or "").strip() or "B001"
        grouped.append((block_key, cleaned[cleaned["block_id"].astype(str).eq(block_key)]))
    return grouped


def _matching_block_rows(rows: pd.DataFrame, block_id: str) -> pd.DataFrame:
    if rows.empty or "block_id" not in rows.columns:
        return rows.iloc[0:0] if not rows.empty else rows
    return rows[rows["block_id"].astype(str).eq(str(block_id))]


def _block_operation_type(group: pd.DataFrame) -> str:
    keys = _operation_key_series(group).drop_duplicates().tolist()
    if len(keys) == 1:
        return str(keys[0])
    return " / ".join(str(key) for key in keys if key)


def _block_status(executable_group: pd.DataFrame, skipped_group: pd.DataFrame) -> str:
    if executable_group.empty and not skipped_group.empty:
        return "失败"
    if not executable_group.empty and not skipped_group.empty:
        return "部分成功"
    if not executable_group.empty:
        return "成功"
    return "失败"


def _first_value(rows: pd.DataFrame, column: str) -> str:
    if rows.empty or column not in rows.columns:
        return ""
    values = rows[column].astype(str).str.strip()
    values = values[values != ""]
    return values.iloc[0] if not values.empty else ""


def _asin_labels(rows: pd.DataFrame) -> list[str]:
    if rows.empty or "keyword" not in rows.columns:
        return []
    labels = []
    for row in rows.to_dict("records"):
        action_type = str(row.get("action_type", ""))
        keyword = _clean_text(row.get("keyword", ""))
        if action_type == "create_product_target" or action_type in PRODUCT_TARGETING_BID_ACTIONS or (action_type == "pause_target" and _looks_like_asin(keyword)):
            asin = _canonical_asin(keyword)
            if asin:
                labels.append(asin)
    return labels


def _keyword_labels(rows: pd.DataFrame) -> list[str]:
    if rows.empty or "keyword" not in rows.columns:
        return []
    labels = []
    for row in rows.to_dict("records"):
        action_type = str(row.get("action_type", ""))
        keyword = _clean_text(row.get("keyword", ""))
        if keyword and action_type in {"create_keyword_target", "update_bid", "add_negative_exact", "add_negative_phrase"}:
            labels.append(keyword)
        elif keyword and action_type == "pause_target" and not _looks_like_asin(keyword):
            labels.append(keyword)
    return labels


def _parse_failure_prechecks(operation_blocks: list[dict[str, str]], actions: pd.DataFrame) -> list[dict[str, object]]:
    if not operation_blocks:
        return []
    parsed_block_ids = set()
    if not actions.empty and "block_id" in actions.columns:
        parsed_block_ids = {str(value) for value in actions["block_id"].astype(str).tolist() if str(value).strip()}
    failures: list[dict[str, object]] = []
    for block in operation_blocks:
        block_id = str(block.get("block_id", ""))
        if block_id in parsed_block_ids:
            continue
        preview = str(block.get("text", ""))[:500].replace("\n", "\\n")
        failures.append(
            {
                "block_id": block_id,
                "block_index": int(str(block.get("block_index", "0")) or 0),
                "block_title": block.get("block_title", "") or "未识别操作",
                "operation_type": "未识别到动作",
                "action_type": "",
                "status": "失败",
                "campaign_name": "",
                "ad_group_name": "",
                "budget": "",
                "bid": "",
                "placement": "",
                "percentage": "",
                "keywords": "",
                "asins": "",
                "keyword_count": 0,
                "asin_count": 0,
                "negative_count": 0,
                "matched_objects": "",
                "unmatched_objects": "",
                "matched_asins": "",
                "unmatched_asins": "",
                "missing_fields": f"未识别到动作。原始操作块前500字：{preview}",
                "suggestion": "请使用“操作：修改关键词竞价”这类模板，并补齐对应字段。",
                "generated_rows": 0,
                "recognized_fields": "",
            }
        )
    return failures


def _suggestion_for_reason(reason: str) -> str:
    text = str(reason or "")
    if not text:
        return ""
    if "竞价" in text:
        return "建议：请补充“竞价：0.6”。"
    if "预算" in text:
        return "建议：请补充“预算：5”。"
    if "SKU" in text:
        return "建议：请补充“SKU：你的SKU”。"
    if "广告组" in text and "多个" in text:
        return "建议：请补充“广告组：广告组名称”。"
    if "广告活动" in text and "找不到" in text:
        return "建议：请重新下载包含该广告活动的 Bulk 表。"
    if "Product Targeting" in text or "ASIN" in text:
        return "建议：请重新下载包含该广告活动和 Product Targeting 的 Bulk 表。"
    if "关键词" in text and "找不到" in text:
        return "建议：请重新下载包含该关键词的 Bulk 表，或确认关键词文本完全一致。"
    return "建议：请检查该操作块的必填字段和原 Bulk 表对象是否存在。"


def _precheck_counts(prechecks: list[dict[str, object]]) -> dict[str, int]:
    total = len(prechecks)
    success = sum(1 for item in prechecks if item.get("status") == "成功")
    partial = sum(1 for item in prechecks if item.get("status") == "部分成功")
    failed = sum(1 for item in prechecks if item.get("status") == "失败")
    return {"total": total, "success": success, "partial": partial, "failed": failed}


def _canonical_action_type(action_type: object, object_type: object = "", keyword: object = "") -> str:
    raw = str(action_type or "").strip()
    if raw in {"create_campaign", "create_ad_group", "create_product_ad", "create_keyword_target"}:
        return "create_keyword_campaign"
    if raw == "create_product_target":
        return "create_product_targeting_campaign"
    if raw == "update_bid":
        return "update_keyword_bid"
    if raw in PRODUCT_TARGETING_BID_ACTIONS:
        return "update_product_targeting_bid"
    if raw == "update_budget":
        return "update_campaign_budget"
    if raw == "pause_target":
        return "pause_target"
    return raw


def _operation_groups(actions: pd.DataFrame) -> list[tuple[str, pd.DataFrame]]:
    if actions.empty:
        return []
    cleaned = _clean_actions(actions)
    keys = _operation_key_series(cleaned)
    grouped: list[tuple[str, pd.DataFrame]] = []
    for key in keys.drop_duplicates().tolist():
        grouped.append((str(key), cleaned[keys == key]))
    return grouped


def _matching_operation_rows(rows: pd.DataFrame, operation_type: str) -> pd.DataFrame:
    if rows.empty or "action_type" not in rows.columns:
        return rows.iloc[0:0] if not rows.empty else rows
    keys = _operation_key_series(rows)
    return rows[keys == operation_type]


def _operation_key_series(rows: pd.DataFrame) -> pd.Series:
    product_campaigns = set()
    if "campaign_name" in rows.columns and "action_type" in rows.columns:
        product_campaigns = {
            _norm_key(name)
            for name in rows.loc[rows["action_type"].astype(str).eq("create_product_target"), "campaign_name"].tolist()
            if _norm_key(name)
        }

    def operation_key(row: pd.Series) -> str:
        action_type = str(row.get("action_type", ""))
        if action_type.startswith("create_"):
            campaign_key = _norm_key(row.get("campaign_name", ""))
            if campaign_key in product_campaigns:
                return "create_product_targeting_campaign"
        return _canonical_action_type(action_type, row.get("object_type", ""), row.get("keyword", ""))

    return rows.apply(operation_key, axis=1)


def _operation_label(operation_type: str) -> str:
    labels = {
        "create_keyword_campaign": "新建关键词广告活动",
        "create_product_targeting_campaign": "新建 ASIN 定投广告活动",
        "update_keyword_bid": "修改关键词竞价",
        "update_product_targeting_bid": "修改 ASIN 定投竞价",
        "update_campaign_budget": "修改广告活动预算",
        "add_negative_exact": "添加否定精准",
        "add_negative_phrase": "添加否定词组",
        "update_placement": "修改广告位百分比",
        "pause_campaign": "暂停广告活动",
        "pause_target": "暂停关键词或 ASIN 投放",
    }
    return labels.get(operation_type, operation_type)


def _count_keyword_rows(rows: pd.DataFrame) -> int:
    if rows.empty or "keyword" not in rows.columns:
        return 0
    count = 0
    for row in rows.to_dict("records"):
        action_type = str(row.get("action_type", ""))
        keyword = _clean_text(row.get("keyword", ""))
        if not keyword:
            continue
        if action_type in {"create_keyword_target", "update_bid"}:
            count += 1
        elif action_type == "pause_target" and not _looks_like_asin(keyword):
            count += 1
    return count


def _count_asin_rows(rows: pd.DataFrame) -> int:
    if rows.empty or "keyword" not in rows.columns:
        return 0
    count = 0
    for row in rows.to_dict("records"):
        action_type = str(row.get("action_type", ""))
        keyword = _clean_text(row.get("keyword", ""))
        if not keyword:
            continue
        if action_type == "create_product_target" or action_type in PRODUCT_TARGETING_BID_ACTIONS:
            count += 1
        elif action_type == "pause_target" and _looks_like_asin(keyword):
            count += 1
    return count


def _count_negative_rows(rows: pd.DataFrame) -> int:
    if rows.empty:
        return 0
    return int(rows["action_type"].isin(["add_negative_exact", "add_negative_phrase"]).sum())


def _object_labels(rows: pd.DataFrame) -> list[str]:
    if rows.empty:
        return []
    labels: list[str] = []
    for row in rows.to_dict("records"):
        action_type = str(row.get("action_type", ""))
        keyword = _clean_text(row.get("keyword", ""))
        campaign_name = _clean_text(row.get("campaign_name", ""))
        ad_group_name = _clean_text(row.get("ad_group_name", ""))
        placement = _clean_text(row.get("placement", ""))
        if action_type in {"create_campaign", "update_budget", "pause_campaign"} and campaign_name:
            labels.append(campaign_name)
        elif action_type == "update_placement":
            labels.append(f"{campaign_name} / {_placement_display(placement)}".strip(" /"))
        elif action_type in PRODUCT_TARGETING_BID_ACTIONS or action_type == "create_product_target" or (action_type == "pause_target" and _looks_like_asin(keyword)):
            labels.append(_canonical_asin(keyword) or keyword)
        elif action_type in {"create_keyword_target", "update_bid", "pause_target", "add_negative_exact", "add_negative_phrase"}:
            labels.append(keyword)
        elif ad_group_name:
            labels.append(ad_group_name)
        elif campaign_name:
            labels.append(campaign_name)
    return labels


def _missing_fields(rows: pd.DataFrame) -> list[str]:
    if rows.empty or "skip_reason" not in rows.columns:
        return []
    fields: list[str] = []
    for reason in rows["skip_reason"].tolist():
        for part in re.split(r"[;；]+", str(reason or "")):
            text = part.strip()
            if not text:
                continue
            if any(token in text for token in ["缺", "必须", "不合法", "填写", "找不到"]):
                fields.append(text)
    return fields


def _recognized_fields(rows: pd.DataFrame) -> str:
    if rows.empty:
        return ""
    names = []
    field_labels = {
        "portfolio_name": "广告组合",
        "campaign_name": "广告活动",
        "ad_group_name": "广告组",
        "sku": "SKU",
        "match_type": "匹配方式",
        "bidding_strategy": "竞价方式",
        "budget": "预算",
        "bid": "竞价",
        "placement": "广告位",
        "percentage": "广告位百分比",
        "state": "状态",
    }
    for field, label in field_labels.items():
        if field in rows.columns and rows[field].astype(str).str.strip().ne("").any():
            names.append(label)
    if _count_keyword_rows(rows):
        names.append("关键词")
    if _count_asin_rows(rows):
        names.append("ASIN")
    if _count_negative_rows(rows):
        names.append("否词")
    return _joined_unique(names)


def _is_nonfatal_skip(errors: list[str]) -> bool:
    text = "；".join(str(error) for error in errors)
    if re.search(r"(^|[；;])\s*缺(?!少 Product Targeting|少商品投放|少广告活动编号|少广告组编号)", text):
        return False
    if any(token in text for token in ["必须", "不合法", "重复", "已存在", "模板缺少"]):
        return False
    return any(token in text for token in ["找不到", "未在上传的 Bulk 表中找到", "不包含", "尚未创建", "匹配到多个", "请填写广告组"])


def _parse_create_ad_section(section: dict[str, str]) -> list[dict[str, object]]:
    fields = _parse_fields(section["body"])
    title_and_body = _source_text(section)
    match_type = _canonical_match_type(fields.get("match_type", "")) or _detect_match_type(title_and_body)
    sku = fields.get("sku", "") or _extract_sku(title_and_body)
    keywords = fields.get("keywords", [])
    if not keywords:
        inline_keyword = _extract_inline_keyword(title_and_body, match_type)
        keywords = [inline_keyword] if inline_keyword else []
    first_keyword = keywords[0] if keywords else ""
    campaign_name = fields.get("campaign_name", "") or _extract_inline_campaign(title_and_body)
    portfolio_name = fields.get("portfolio_name", "")
    portfolio_id = fields.get("portfolio_id", "")
    if not campaign_name and sku and first_keyword:
        campaign_name = f"SP-{sku}-{_slug_keyword(first_keyword)}-{_campaign_suffix(match_type)}"
    ad_group_name = fields.get("ad_group_name", "") or campaign_name
    state = _canonical_state(fields.get("state", "")) or "enabled"
    bidding_strategy = _canonical_bidding_strategy(fields.get("bidding_strategy", ""))
    budget = fields.get("budget", "") or _extract_number_after(title_and_body, "预算")
    bid = fields.get("bid", "") or _extract_number_after(title_and_body, "竞价") or _extract_number_after(title_and_body, "出价")

    rows = [
        _action_row("create_campaign", "campaign", campaign_name, "", sku, "", match_type, bidding_strategy, budget, "", "", state, "新建关键词广告活动", title_and_body, portfolio_name=portfolio_name, portfolio_id=portfolio_id),
        _action_row("create_ad_group", "ad_group", campaign_name, ad_group_name, sku, "", match_type, bidding_strategy, "", bid, "", state, "新建关键词广告活动", title_and_body, portfolio_name=portfolio_name, portfolio_id=portfolio_id),
        _action_row("create_product_ad", "product_ad", campaign_name, ad_group_name, sku, "", match_type, bidding_strategy, "", "", "", state, "新建关键词广告活动", title_and_body, portfolio_name=portfolio_name, portfolio_id=portfolio_id),
    ]
    if keywords:
        for keyword in keywords:
            rows.append(_action_row("create_keyword_target", "targeting", campaign_name, ad_group_name, sku, keyword, match_type, bidding_strategy, "", bid, "", state, "新建关键词广告活动", title_and_body, portfolio_name=portfolio_name, portfolio_id=portfolio_id))
    else:
        rows.append(_action_row("create_keyword_target", "targeting", campaign_name, ad_group_name, sku, "", match_type, bidding_strategy, "", bid, "", state, "新建关键词广告活动", title_and_body, portfolio_name=portfolio_name, portfolio_id=portfolio_id))
    return rows


def _parse_create_product_targeting_campaign_section(section: dict[str, str]) -> list[dict[str, object]]:
    fields = _parse_fields(section["body"])
    title_and_body = _source_text(section)
    sku = fields.get("sku", "") or _extract_sku(title_and_body)
    asins = fields.get("asins", []) or _extract_asins(title_and_body)
    first_asin = asins[0] if asins else ""
    campaign_name = fields.get("campaign_name", "") or _extract_inline_campaign(title_and_body)
    if not campaign_name and sku and first_asin:
        campaign_name = f"SP-{sku}-{_canonical_asin(first_asin)}-ASIN"
    ad_group_name = fields.get("ad_group_name", "") or campaign_name
    state = _canonical_state(fields.get("state", "")) or "enabled"
    bidding_strategy = _canonical_bidding_strategy(fields.get("bidding_strategy", ""))
    budget = fields.get("budget", "") or _extract_number_after(title_and_body, "预算")
    bid = fields.get("bid", "") or _extract_number_after(title_and_body, "竞价") or _extract_number_after(title_and_body, "出价")

    rows = [
        _action_row("create_campaign", "campaign", campaign_name, "", sku, "", "", bidding_strategy, budget, "", "", state, "新建ASIN定投广告活动", title_and_body),
        _action_row("create_ad_group", "ad_group", campaign_name, ad_group_name, sku, "", "", bidding_strategy, "", bid, "", state, "新建ASIN定投广告活动", title_and_body),
        _action_row("create_product_ad", "product_ad", campaign_name, ad_group_name, sku, "", "", bidding_strategy, "", "", "", state, "新建ASIN定投广告活动", title_and_body),
    ]
    if asins:
        for asin in asins:
            rows.append(_action_row("create_product_target", "product_targeting", campaign_name, ad_group_name, sku, _canonical_asin(asin), "", bidding_strategy, "", bid, "", state, "新建ASIN定投广告活动", title_and_body))
    else:
        rows.append(_action_row("create_product_target", "product_targeting", campaign_name, ad_group_name, sku, "", "", bidding_strategy, "", bid, "", state, "新建ASIN定投广告活动", title_and_body))
    return rows


def _parse_negative_section(section: dict[str, str], action_type: str) -> list[dict[str, object]]:
    fields = _parse_fields(section["body"])
    source = _source_text(section)
    campaign_name = fields.get("campaign_name", "") or _extract_inline_campaign(source)
    ad_group_name = fields.get("ad_group_name", "")
    keywords = fields.get("negative_keywords", []) or fields.get("keywords", []) or _extract_inline_negative_keywords(source, action_type)
    if not _has_optional_action_content(campaign_name, ad_group_name, keywords):
        return []
    negative_type = "negative exact" if action_type == "add_negative_exact" else "negative phrase"
    if not keywords:
        keywords = [""]
    return [
        _action_row(action_type, "negative_keyword", campaign_name, ad_group_name, "", keyword, "", "", "", "", negative_type, "enabled", "添加否定词", source)
        for keyword in keywords
    ]


def _parse_budget_section(section: dict[str, str]) -> dict[str, object] | None:
    source = _source_text(section)
    fields = _parse_fields(section["body"])
    campaign_name = fields.get("campaign_name", "") or _extract_inline_campaign(source)
    budget = fields.get("budget", "") or _extract_number_after(source, "预算")
    if not _has_optional_action_content(campaign_name, budget):
        return None
    return _action_row("update_budget", "campaign", campaign_name, "", "", "", "", "", budget, "", "", "enabled", "修改广告活动预算", source)


def _parse_bid_section(section: dict[str, str]) -> list[dict[str, object]]:
    source = _source_text(section)
    fields = _parse_fields(section["body"])
    campaign_name = fields.get("campaign_name", "") or _extract_inline_campaign(source)
    ad_group_name = fields.get("ad_group_name", "")
    bid_pairs = _extract_keyword_bid_pairs(source)
    if bid_pairs:
        return [
            _action_row("update_bid", "targeting", campaign_name, ad_group_name, "", keyword, "", "", "", bid, "", "enabled", "修改关键词竞价", source)
            for keyword, bid in bid_pairs
            if _has_optional_action_content(campaign_name, ad_group_name, keyword, bid)
        ]
    keywords = fields.get("keywords", [])
    bid = fields.get("bid", "") or _extract_number_after(source, "竞价") or _extract_number_after(source, "出价")
    rows = []
    for item in keywords:
        pair = re.match(r"^(.+?)\s*(?:=|:|：|->|→)\s*([$￥]?\s*[0-9,]+(?:\.[0-9]+)?\s*(?:美金|美元)?)\s*$", str(item).strip())
        keyword = _clean_text(pair.group(1)) if pair else _clean_text(item)
        item_bid = _clean_number_text(pair.group(2)) if pair else bid
        if _has_optional_action_content(campaign_name, ad_group_name, keyword, item_bid):
            rows.append(_action_row("update_bid", "targeting", campaign_name, ad_group_name, "", keyword, "", "", "", item_bid, "", "enabled", "修改关键词竞价", source))
    if rows:
        return rows
    keyword = fields.get("keyword", "") or _extract_keyword_before_bid(source)
    if not _has_optional_action_content(campaign_name, ad_group_name, keyword, bid):
        return []
    return [_action_row("update_bid", "targeting", campaign_name, ad_group_name, "", keyword, "", "", "", bid, "", "enabled", "修改关键词竞价", source)]


def _parse_product_target_bid_section(section: dict[str, str]) -> list[dict[str, object]]:
    source = _source_text(section)
    fields = _parse_fields(section["body"])
    campaign_name = fields.get("campaign_name", "") or _extract_inline_campaign(source)
    ad_group_name = fields.get("ad_group_name", "")
    bid_pairs = _extract_keyword_bid_pairs(source)
    if bid_pairs:
        return [
            _action_row("update_product_targeting_bid", "product_targeting", campaign_name, ad_group_name, "", _canonical_asin(target), "", "", "", bid, "", "enabled", "修改ASIN定投竞价", source)
            for target, bid in bid_pairs
            if _has_optional_action_content(campaign_name, ad_group_name, target, bid)
        ]
    asins = fields.get("asins", []) or _extract_asins(source)
    bid = fields.get("bid", "") or _extract_number_after(source, "竞价") or _extract_number_after(source, "出价")
    rows = []
    for item in asins:
        pair = re.match(r"^(.+?)\s*(?:=|:|：|->|→)\s*([$￥]?\s*[0-9,]+(?:\.[0-9]+)?\s*(?:美金|美元)?)\s*$", str(item).strip())
        asin = _canonical_asin(pair.group(1) if pair else item)
        item_bid = _clean_number_text(pair.group(2)) if pair else bid
        if _has_optional_action_content(campaign_name, ad_group_name, asin, item_bid):
            rows.append(_action_row("update_product_targeting_bid", "product_targeting", campaign_name, ad_group_name, "", asin, "", "", "", item_bid, "", "enabled", "修改ASIN定投竞价", source))
    if rows:
        return rows
    if not _has_optional_action_content(campaign_name, ad_group_name, asins, bid):
        return []
    return [
        _action_row("update_product_targeting_bid", "product_targeting", campaign_name, ad_group_name, "", _canonical_asin(asin), "", "", "", bid, "", "enabled", "修改ASIN定投竞价", source)
        for asin in asins
        if _canonical_asin(asin)
    ]


def _parse_placement_section(section: dict[str, str]) -> list[dict[str, object]]:
    source = _source_text(section)
    fields = _parse_fields(section["body"])
    campaign_name = fields.get("campaign_name", "") or _extract_inline_campaign(source)
    placements = _extract_placements(source)
    rows = []
    for placement, percentage in placements:
        row = _action_row("update_placement", "placement", campaign_name, "", "", "", "", "", "", "", "", "enabled", "修改广告位百分比", source)
        row["placement"] = placement
        row["percentage"] = _clean_number_text(percentage)
        rows.append(row)
    return rows


def _parse_pause_campaign_section(section: dict[str, str]) -> dict[str, object] | None:
    source = _source_text(section)
    fields = _parse_fields(section["body"])
    campaign_name = fields.get("campaign_name", "") or _extract_inline_campaign(source)
    if not _has_optional_action_content(campaign_name):
        return None
    return _action_row("pause_campaign", "campaign", campaign_name, "", "", "", "", "", "", "", "", "paused", "暂停广告活动", source)


def _parse_pause_target_section(section: dict[str, str]) -> list[dict[str, object]]:
    source = _source_text(section)
    fields = _parse_fields(section["body"])
    campaign_name = fields.get("campaign_name", "") or _extract_inline_campaign(source)
    ad_group_name = fields.get("ad_group_name", "")
    asins = fields.get("asins", []) or _extract_asins(source)
    rows = [
        _action_row("pause_target", "product_targeting", campaign_name, ad_group_name, "", _canonical_asin(asin), "", "", "", "", "", "paused", "暂停ASIN投放", source)
        for asin in asins
        if _canonical_asin(asin)
    ]
    keywords = fields.get("keywords", []) or _extract_pause_keywords(source)
    rows.extend(
        _action_row("pause_target", "targeting", campaign_name, ad_group_name, "", keyword, "", "", "", "", "", "paused", "暂停关键词投放", source)
        for keyword in keywords
    )
    return rows


def _has_optional_action_content(*values: object) -> bool:
    for value in values:
        if isinstance(value, list):
            if any(str(item).strip() for item in value):
                return True
        elif str(value or "").strip():
            return True
    return False


def _action_errors(
    action: dict[str, object],
    template: BulkTemplate | None,
    indexes: BulkIndexes,
    create_campaigns: set[str],
    create_ad_groups: set[tuple[str, str]],
) -> list[str]:
    errors = []
    action_type = str(action.get("action_type", ""))
    campaign_id = str(action.get("campaign_id", "")).strip()
    ad_group_id = str(action.get("ad_group_id", "")).strip()
    campaign_name = str(action.get("campaign_name", "")).strip()
    ad_group_name = str(action.get("ad_group_name", "")).strip()
    sku = str(action.get("sku", "")).strip()
    keyword = str(action.get("keyword", "")).strip()
    match_type = str(action.get("match_type", "")).strip()
    bidding_strategy = str(action.get("bidding_strategy", "")).strip()
    portfolio_name = str(action.get("portfolio_name", "")).strip()
    portfolio_id = str(action.get("portfolio_id", "")).strip()
    budget = _num(action.get("budget", ""))
    bid = _num(action.get("bid", ""))

    if template is None:
        errors.append("缺少可用 Bulk 模板")
    if action_type not in SUPPORTED_ACTIONS:
        errors.append("不支持的动作类型")
    if action_type == "create_campaign":
        if not campaign_id:
            errors.append("create_campaign 行缺少 Campaign ID")
        if not campaign_name:
            errors.append("缺广告活动名称")
        elif _campaign_exists(indexes, campaign_name):
            errors.append("广告活动名称已存在")
        if portfolio_name and not portfolio_id:
            errors.append("找不到广告组合编号，请确认原始 Bulk 表包含该广告组合。")
        if not bidding_strategy:
            errors.append("缺竞价方式")
        if budget <= 0:
            errors.append("缺预算或预算必须大于 0")
    elif action_type == "create_ad_group":
        if not campaign_id or not ad_group_id:
            errors.append("create_ad_group 行缺少 Campaign ID 或 Ad Group ID")
        if not campaign_name:
            errors.append("缺广告活动名称")
        if not ad_group_name:
            errors.append("缺广告组名称")
        if bid <= 0:
            errors.append("缺竞价或竞价必须大于 0")
    elif action_type == "create_product_ad":
        if not campaign_id or not ad_group_id:
            errors.append("create_product_ad 行缺少 Campaign ID 或 Ad Group ID")
        if not campaign_name:
            errors.append("缺广告活动名称")
        if not ad_group_name:
            errors.append("缺广告组名称")
        if not sku:
            errors.append("缺 SKU")
    elif action_type == "create_keyword_target":
        if not campaign_id or not ad_group_id:
            errors.append("create_keyword_target 行缺少 Campaign ID 或 Ad Group ID")
        if not campaign_name:
            errors.append("缺广告活动名称")
        if not ad_group_name:
            errors.append("缺广告组名称")
        if not keyword:
            errors.append("缺关键词")
        if match_type not in MATCH_TYPES:
            errors.append("匹配方式不合法")
        if bid <= 0:
            errors.append("缺竞价或竞价必须大于 0")
        if campaign_name not in create_campaigns and not _campaign_exists(indexes, campaign_name):
            errors.append("找不到广告活动")
        if (_norm_key(campaign_name), _norm_key(ad_group_name)) not in create_ad_groups and not _ad_group_exists(indexes, campaign_name, ad_group_name):
            errors.append("找不到广告组")
    elif action_type == "create_product_target":
        if not campaign_id or not ad_group_id:
            errors.append("create_product_target 行缺少 Campaign ID 或 Ad Group ID")
        if not campaign_name:
            errors.append("缺广告活动名称")
        if not ad_group_name:
            errors.append("缺广告组名称")
        if not keyword:
            errors.append("缺 ASIN")
        if bid <= 0:
            errors.append("缺竞价或竞价必须大于 0")
        if campaign_name not in create_campaigns and not _campaign_exists(indexes, campaign_name):
            errors.append("找不到广告活动")
        if (_norm_key(campaign_name), _norm_key(ad_group_name)) not in create_ad_groups and not _ad_group_exists(indexes, campaign_name, ad_group_name):
            errors.append("找不到广告组")
    elif action_type == "update_budget":
        if not campaign_name:
            errors.append("缺广告活动名称")
        elif not _campaign_exists(indexes, campaign_name):
            errors.append("找不到广告活动")
        elif not campaign_id:
            errors.append("找不到广告活动编号，请重新下载包含该广告活动的 Bulk 表。")
        if budget <= 0:
            errors.append("缺预算或预算必须大于 0")
    elif action_type == "update_bid":
        matches = _find_target(indexes, campaign_name, ad_group_name, keyword)
        if not campaign_name:
            errors.append("缺广告活动名称")
        elif not _campaign_exists(indexes, campaign_name):
            errors.append("找不到广告活动")
        elif keyword and not matches.empty and (not campaign_id or not ad_group_id):
            errors.append("找不到广告活动编号或广告组编号，请重新下载包含该广告活动的 Bulk 表。")
        if keyword and not matches.empty and template is not None and "keyword_id" in template.header_map and not action.get("keyword_id", "") and not action.get("entity_id", "") and not action.get("product_target_id", ""):
            errors.append("找不到关键词编号，请重新下载包含该关键词的 Bulk 表。")
        if not keyword:
            errors.append("缺关键词")
        if bid <= 0:
            errors.append("缺竞价或竞价必须大于 0")
        if keyword and matches.empty:
            errors.append("找不到关键词")
        elif keyword and len(matches) > 1 and not ad_group_name:
            errors.append("关键词匹配到多个广告组，请填写广告组")
    elif action_type in PRODUCT_TARGETING_BID_ACTIONS:
        if not campaign_name:
            errors.append("缺广告活动名称")
        if not keyword:
            errors.append("缺 ASIN")
        if bid <= 0:
            errors.append("缺竞价或竞价必须大于 0")
        matches = _find_product_target(indexes, campaign_name, ad_group_name, keyword)
        if keyword and matches.empty:
            errors.append(_product_target_not_found_reason(indexes))
        elif keyword and len(matches) > 1 and not ad_group_name:
            errors.append("ASIN 定投匹配到多个广告组，请填写广告组")
        elif keyword and (not campaign_id or not ad_group_id or (not action.get("product_target_id", "") and not action.get("entity_id", ""))):
            errors.append("找到 ASIN 定投，但原 Bulk 表缺少广告活动编号、广告组编号或商品投放编号，请重新下载包含 Product Targeting ID 的 Bulk 表。")
    elif action_type == "update_placement":
        if not campaign_name:
            errors.append("缺广告活动名称")
        elif not _campaign_exists(indexes, campaign_name):
            errors.append("找不到广告活动")
        elif not campaign_id:
            errors.append("找不到广告活动编号，请重新下载包含该广告活动的 Bulk 表。")
        if not action.get("placement", ""):
            errors.append("缺广告位")
        if str(action.get("percentage", "")).strip() == "" or _num(action.get("percentage", "")) < 0:
            errors.append("缺广告位百分比或百分比必须大于等于 0")
    elif action_type == "pause_campaign":
        if not campaign_name:
            errors.append("缺广告活动名称")
        elif not _campaign_exists(indexes, campaign_name):
            errors.append("找不到广告活动")
        elif not campaign_id:
            errors.append("找不到广告活动编号，请重新下载包含该广告活动的 Bulk 表。")
    elif action_type == "pause_target":
        if not campaign_name:
            errors.append("缺广告活动名称")
        if not keyword:
            errors.append("缺关键词或 ASIN")
        if keyword and _looks_like_asin(keyword):
            matches = _find_product_target(indexes, campaign_name, ad_group_name, keyword)
            if matches.empty:
                errors.append(_product_target_not_found_reason(indexes))
            elif len(matches) > 1 and not ad_group_name:
                errors.append("ASIN 投放匹配到多个广告组，请填写广告组")
            elif not campaign_id or not ad_group_id or (not action.get("product_target_id", "") and not action.get("entity_id", "")):
                errors.append("找到 ASIN 投放，但原 Bulk 表缺少广告活动编号、广告组编号或商品投放编号。")
        else:
            matches = _find_target(indexes, campaign_name, ad_group_name, keyword)
            if keyword and matches.empty:
                errors.append("找不到关键词")
            elif keyword and len(matches) > 1 and not ad_group_name:
                errors.append("关键词匹配到多个广告组，请填写广告组")
            elif not campaign_id or not ad_group_id or (not action.get("keyword_id", "") and not action.get("entity_id", "")):
                errors.append("找到关键词，但原 Bulk 表缺少广告活动编号、广告组编号或关键词编号。")
    elif action_type in {"add_negative_exact", "add_negative_phrase"}:
        if not campaign_name:
            errors.append("缺广告活动名称")
        elif not _campaign_exists(indexes, campaign_name):
            errors.append("找不到广告活动")
        elif not campaign_id:
            errors.append("找不到广告活动编号，请重新下载包含该广告活动的 Bulk 表。")
        if ad_group_name and not ad_group_id:
            errors.append("找不到广告组编号，请重新下载包含该广告活动的 Bulk 表。")
        if not keyword:
            errors.append("缺关键词")
        negative_type = "negative exact" if action_type == "add_negative_exact" else "negative phrase"
        if keyword and _negative_exists(indexes, campaign_name, ad_group_name, keyword, negative_type):
            errors.append("否定词已存在")
    return errors


def _new_campaign_structure_errors(actions: pd.DataFrame) -> dict[str, list[str]]:
    errors = {}
    for campaign_name in actions.loc[actions["action_type"] == "create_campaign", "campaign_name"].unique():
        group = actions[actions["campaign_name"] == campaign_name]
        campaign_errors = []
        if not (group["action_type"] == "create_ad_group").any():
            campaign_errors.append("新建广告缺 Ad Group 行")
        if not (group["action_type"] == "create_product_ad").any():
            campaign_errors.append("新建广告缺 Product Ad 行")
        has_keyword_target = (group["action_type"] == "create_keyword_target").any()
        has_product_target = (group["action_type"] == "create_product_target").any()
        if not has_keyword_target and not has_product_target:
            campaign_errors.append("新建广告缺 Keyword 或 Product Targeting 行")
        if campaign_errors:
            errors[str(campaign_name)] = campaign_errors
    return errors


def _required_column_validation(template: BulkTemplate, actions: pd.DataFrame) -> list[dict[str, object]]:
    required = set(BASE_REQUIRED_BULK_FIELDS)
    needs_portfolio = (
        "portfolio_id" in actions.columns
        and actions["portfolio_id"].astype(str).str.strip().ne("").any()
    ) or (
        "portfolio_name" in actions.columns
        and actions["portfolio_name"].astype(str).str.strip().ne("").any()
    )
    if needs_portfolio:
        required.add("portfolio_id")
    for action_type in actions["action_type"].unique():
        if action_type == "create_campaign":
            required.update(["campaign_id", "daily_budget", "targeting_type", "bidding_strategy"])
        elif action_type == "create_ad_group":
            required.update(["campaign_id", "ad_group_id", "ad_group_name", "ad_group_default_bid"])
        elif action_type == "create_product_ad":
            required.update(["campaign_id", "ad_group_id", "ad_group_name", "sku"])
        elif action_type in {"create_keyword_target", "update_bid"}:
            required.update(["campaign_id", "ad_group_id", "ad_group_name", "match_type", "bid"])
            if "keyword_text" not in template.header_map and "product_targeting_expression" not in template.header_map:
                required.add("keyword_text")
        elif action_type == "create_product_target":
            required.update(["campaign_id", "ad_group_id", "ad_group_name", "bid"])
            if "product_targeting_expression" not in template.header_map and "keyword_text" not in template.header_map:
                required.add("product_targeting_expression")
        elif action_type in PRODUCT_TARGETING_BID_ACTIONS:
            required.update(["campaign_id", "ad_group_id", "ad_group_name", "bid"])
            if "product_targeting_expression" not in template.header_map and "keyword_text" not in template.header_map:
                required.add("product_targeting_expression")
        elif action_type == "update_placement":
            required.update(["campaign_id", "placement", "percentage"])
        elif action_type == "pause_campaign":
            required.update(["campaign_id", "state"])
        elif action_type == "pause_target":
            required.update(["campaign_id", "ad_group_id", "state"])
        elif action_type in {"add_negative_exact", "add_negative_phrase"}:
            required.update(["campaign_id", "keyword_text", "match_type"])
        elif action_type == "update_budget":
            required.update(["campaign_id", "daily_budget"])
    missing = sorted(field for field in required if field not in template.header_map)
    if missing:
        if "portfolio_id" in missing:
            return [_validation_row("Bulk 列校验", "error", "模板缺少广告组合编号列，无法设置广告组合。")]
        if {"entity_id", "campaign_id", "ad_group_id"} & set(missing):
            return [_validation_row("父实体编号校验", "error", "子级行缺少父实体编号，请检查 Campaign ID / Ad Group ID 是否已生成。")]
        return [_validation_row("Bulk 列校验", "error", f"模板缺少必要列：{', '.join(missing)}")]
    return [_validation_row("Bulk 列校验", "ok", "模板列满足当前动作")]


def _bulk_id_validation(template: BulkTemplate, actions: pd.DataFrame) -> list[dict[str, object]]:
    errors: list[dict[str, object]] = []
    for action in actions.to_dict("records"):
        action_type = str(action.get("action_type", ""))
        entity_id = str(action.get("entity_id", "")).strip()
        campaign_id = str(action.get("campaign_id", "")).strip()
        ad_group_id = str(action.get("ad_group_id", "")).strip()
        ad_id = str(action.get("ad_id", "")).strip()
        keyword_id = str(action.get("keyword_id", "")).strip()
        product_target_id = str(action.get("product_target_id", "")).strip()
        missing = False
        if action_type == "create_campaign" and not campaign_id:
            missing = True
        elif action_type in {"create_ad_group", "create_product_ad", "create_keyword_target", "create_product_target"} and (not campaign_id or not ad_group_id):
            missing = True
        if "entity_id" in template.header_map and not entity_id:
            missing = True
        if action_type == "create_product_ad" and "ad_id" in template.header_map and not ad_id:
            missing = True
        if action_type in {"create_keyword_target", "add_negative_exact", "add_negative_phrase"} and "keyword_id" in template.header_map and not keyword_id and not entity_id and not product_target_id:
            missing = True
        if action_type == "create_product_target" and "product_target_id" in template.header_map and not product_target_id and not entity_id:
            missing = True
        if action_type == "update_bid" and "keyword_id" in template.header_map and not keyword_id and not entity_id and not product_target_id:
            missing = True
        if action_type in PRODUCT_TARGETING_BID_ACTIONS and "product_target_id" in template.header_map and not product_target_id and not entity_id:
            missing = True
        if missing:
            errors.append(
                _validation_row(
                    "父实体编号校验",
                    "error",
                    "子级行缺少父实体编号，请检查 Campaign ID / Ad Group ID 是否已生成。",
                    str(action.get("action_id", "")),
                )
            )
    if errors:
        return errors
    if _has_create_child_actions(actions):
        return [_validation_row("父实体编号校验", "ok", "新建广告父子行已写入临时 Campaign ID / Ad Group ID")]
    return []


def _generated_bulk_value_validation(template: BulkTemplate, actions: pd.DataFrame) -> list[dict[str, object]]:
    errors: list[dict[str, object]] = []
    for action in actions.to_dict("records"):
        values = _bulk_values(action, template)
        for field in CONTROLLED_BULK_FIELDS:
            value = values.get(field, "")
            if value in ("", None):
                continue
            value_text = str(value).strip()
            if _has_cjk(value_text):
                column_name = _header_name(template, field)
                errors.append(
                    _validation_row(
                        "Bulk 枚举值校验",
                        "error",
                        f"Bulk 枚举值非法：列“{column_name}”的值“{value_text}”无效，写入 Bulk 时必须使用 Amazon 允许的英文值或上传模板中的原始值。",
                        str(action.get("action_id", "")),
                    )
                )
    if errors:
        return errors
    return [_validation_row("Bulk 枚举值校验", "ok", "Bulk 枚举值使用英文标准值或模板原值")]


def _has_create_child_actions(actions: pd.DataFrame) -> bool:
    return actions["action_type"].isin(["create_ad_group", "create_product_ad", "create_keyword_target", "create_product_target"]).any()


def _bulk_values(action: dict[str, object], template: BulkTemplate) -> dict[str, object]:
    action_type = str(action["action_type"])
    values = {
        "product": _label(template, "product", "sponsored_products"),
        "operation": _label(template, "operation", "update" if action_type in BULK_UPDATE_ACTIONS else "create"),
        "state": _label(template, "state", action.get("state", "enabled")),
        "entity_id": action.get("entity_id", ""),
        "campaign_id": action.get("campaign_id", ""),
        "ad_group_id": action.get("ad_group_id", ""),
        "ad_id": action.get("ad_id", ""),
        "keyword_id": action.get("keyword_id", ""),
        "product_target_id": action.get("product_target_id", ""),
        "campaign_name": action.get("campaign_name", ""),
        "ad_group_name": action.get("ad_group_name", ""),
    }
    if action_type == "create_campaign":
        values.update({"entity": _label(template, "entity", "campaign"), "portfolio_id": action.get("portfolio_id", ""), "daily_budget": action.get("budget", ""), "targeting_type": _label(template, "targeting_type", "manual"), "bidding_strategy": _label(template, "bidding_strategy", action.get("bidding_strategy", ""))})
    elif action_type == "create_ad_group":
        values.update({"entity": _label(template, "entity", "ad_group"), "ad_group_default_bid": action.get("bid", "")})
    elif action_type == "create_product_ad":
        values.update({"entity": _label(template, "entity", "product_ad"), "sku": action.get("sku", "")})
    elif action_type == "create_keyword_target":
        keyword_field, keyword_value = _keyword_bulk_value(template, action.get("keyword", ""))
        values.update({"entity": _label(template, "entity", "keyword_target"), keyword_field: keyword_value, "match_type": _label(template, "match_type", action.get("match_type", "")), "bid": action.get("bid", "")})
    elif action_type == "create_product_target":
        target_field, target_value = _product_target_bulk_value(template, action.get("keyword", ""))
        values.update({"entity": _label(template, "entity", "product_target"), target_field: target_value, "bid": action.get("bid", "")})
    elif action_type == "update_budget":
        values.update({"entity": _label(template, "entity", "campaign"), "daily_budget": action.get("budget", "")})
    elif action_type == "update_bid":
        keyword_field, keyword_value = _keyword_bulk_value(template, action.get("keyword", ""))
        values.update({"entity": _label(template, "entity", "keyword_target"), keyword_field: keyword_value, "match_type": _label(template, "match_type", action.get("match_type", "")), "bid": action.get("bid", "")})
    elif action_type in PRODUCT_TARGETING_BID_ACTIONS:
        target_field, target_value = _product_target_bulk_value(template, action.get("keyword", ""))
        values.update({"entity": _label(template, "entity", "product_target"), target_field: target_value, "bid": action.get("bid", "")})
    elif action_type == "update_placement":
        values.update({"entity": _label(template, "entity", "placement"), "placement": _placement_label(template, action.get("placement", "")), "percentage": action.get("percentage", "")})
    elif action_type == "pause_campaign":
        values.update({"entity": _label(template, "entity", "campaign"), "state": _label(template, "state", "paused")})
    elif action_type == "pause_target":
        values.update({"entity": _label(template, "entity", "product_target" if _looks_like_asin(action.get("keyword", "")) else "keyword_target"), "state": _label(template, "state", "paused")})
    elif action_type in {"add_negative_exact", "add_negative_phrase"}:
        entity_key = "negative_keyword" if action.get("ad_group_name", "") else "campaign_negative_keyword"
        values.update({"entity": _label(template, "entity", entity_key), "keyword_text": action.get("keyword", ""), "match_type": _label(template, "match_type", action.get("negative_type", ""))})
    return {key: value for key, value in values.items() if value not in ("", None)}


def _copy_update_values(action: dict[str, object], template: BulkTemplate) -> dict[str, object]:
    action_type = str(action.get("action_type", ""))
    values = {"operation": _label(template, "operation", "update")}
    if action_type == "update_budget":
        values["daily_budget"] = action.get("budget", "")
    elif action_type in {"update_bid"} | PRODUCT_TARGETING_BID_ACTIONS:
        values["bid"] = action.get("bid", "")
    elif action_type in {"pause_campaign", "pause_target"}:
        values["state"] = _label(template, "state", "paused")
    return values


def _controlled_update_values(action: dict[str, object], template: BulkTemplate) -> dict[str, object]:
    values: dict[str, object] = {
        "product": _label(template, "product", "sponsored_products"),
        "entity": _label(template, "entity", _entity_key_for_action(action)),
    }
    source_row = _source_row_for_action(template, action)
    if source_row is None:
        return {key: value for key, value in values.items() if value not in ("", None)}

    for field, group, canonical_func in [
        ("targeting_type", "targeting_type", _canonical_targeting_type),
        ("state", "state", _canonical_state),
        ("match_type", "match_type", _canonical_match_type),
        ("bidding_strategy", "bidding_strategy", _canonical_bidding_strategy),
    ]:
        raw_value = _row_value(source_row, template.header_map, field)
        canonical = canonical_func(raw_value)
        if canonical:
            values[field] = _label(template, group, canonical)
    return {key: value for key, value in values.items() if value not in ("", None)}


def _entity_key_for_action(action: dict[str, object]) -> str:
    action_type = str(action.get("action_type", ""))
    if action_type in {"update_budget", "pause_campaign"}:
        return "campaign"
    if action_type in PRODUCT_TARGETING_BID_ACTIONS:
        return "product_target"
    if action_type == "pause_target" and _looks_like_asin(action.get("keyword", "")):
        return "product_target"
    if action_type in {"update_bid", "pause_target"}:
        return "keyword_target"
    return str(action.get("object_type", ""))


def _source_row_for_action(template: BulkTemplate, action: dict[str, object]) -> pd.Series | None:
    try:
        row_index = int(float(str(action.get("source_row_index", "")).strip()))
    except ValueError:
        return None
    if row_index < 0 or row_index >= len(template.data):
        return None
    return template.data.iloc[row_index]


def _source_row_values(template: BulkTemplate, action: dict[str, object]) -> list[object]:
    row = _source_row_for_action(template, action)
    if row is None:
        return [None for _ in template.headers]
    return [None if pd.isna(value) else value for value in row.tolist()[: len(template.headers)]]


def _keyword_bulk_value(template: BulkTemplate, keyword: object) -> tuple[str, str]:
    keyword_text = _clean_text(keyword)
    if "keyword_text" in template.header_map:
        return "keyword_text", keyword_text
    return "product_targeting_expression", f'keyword="{keyword_text}"'


def _asin_expression(value: object) -> str:
    return f'asin="{_canonical_asin(value)}"'


def _product_target_bulk_value(template: BulkTemplate, asin: object) -> tuple[str, str]:
    if "product_targeting_expression" in template.header_map:
        return "product_targeting_expression", _asin_expression(asin)
    return "keyword_text", _asin_expression(asin)


def _split_sections(text: str) -> list[dict[str, str]]:
    normalized = text.replace("\r\n", "\n")
    lines = normalized.split("\n")
    sections: list[dict[str, str]] = []
    title = ""
    body: list[str] = []
    for line in lines:
        stripped = line.strip()
        if _line_starts_section(stripped):
            if title or body:
                sections.append({"title": title, "body": "\n".join(body).strip()})
            title = stripped.rstrip("：:")
            body = []
        else:
            body.append(line)
    if title or body:
        sections.append({"title": title, "body": "\n".join(body).strip()})
    if len(sections) == 1 and not sections[0]["title"]:
        return [{"title": part.strip(), "body": part.strip()} for part in re.split(r"[。；;]+", text) if part.strip()]
    return sections


def _line_starts_section(line: str) -> bool:
    if (re.search(r"\d", line) or re.search(r"[:：]", line)) and any(line.startswith(token) for token in ["首页首位", "Top of Search", "商品页面", "Product Pages", "其余位置", "Rest of Search"]):
        return False
    triggers = [
        "新建广告",
        "创建广告",
        "开广告",
        "开精准",
        "开词组",
        "开广泛",
        "新建关键词广告",
        "新建ASIN定投",
        "开ASIN定投",
        "投这个ASIN",
        "投放竞品ASIN",
        "商品定投",
        "开个精准关键词集合",
        "新建精准关键词集合",
        "开词组广告",
        "开广泛广告",
        "添加否定精准",
        "添加否定词组",
        "否定精准",
        "否定词组",
        "可选调整项",
        "修改预算",
        "预算改成",
        "预算加到",
        "预算降低到",
        "修改关键词竞价",
        "改关键词竞价",
        "关键词竞价改成",
        "这些词加价",
        "这些词降价",
        "修改ASIN定投竞价",
        "修改ASIN竞价",
        "ASIN定投竞价",
        "商品投放竞价",
        "修改商品投放竞价",
        "修改商品定投竞价",
        "Product Targeting Bid",
        "修改竞价",
        "首页首位",
        "Top of Search",
        "商品页面",
        "Product Pages",
        "其余位置",
        "Rest of Search",
        "placement",
        "广告位百分比",
        "修改广告位",
        "暂停广告活动",
        "关闭广告活动",
        "campaign pause",
        "暂停关键词",
        "暂停这个词",
        "暂停ASIN",
        "暂停投放",
        "关闭这个投放",
    ]
    return bool(line) and any(line.startswith(trigger) for trigger in triggers)


def _section_kind(title: str, body: str) -> str:
    text = f"{title}\n{body}"
    lowered = text.lower()
    if any(token in text for token in ["新建ASIN定投", "开ASIN定投", "投这个ASIN", "投放竞品ASIN", "商品定投"]):
        return "create_product_targeting_campaign"
    if any(token in text for token in ["新建广告", "创建广告", "开广告", "开精准", "开词组", "开广泛", "新建关键词广告", "开个精准关键词集合", "新建精准关键词集合", "开词组广告", "开广泛广告"]):
        return "create_keyword_campaign"
    if any(token in text for token in ["添加否定精准", "否定精准", "否精准", "精准否定"]) or "negative exact" in lowered:
        return "add_negative_exact"
    if any(token in text for token in ["添加否定词组", "否定词组", "否词组", "词组否定"]) or "negative phrase" in lowered:
        return "add_negative_phrase"
    if any(token in text for token in ["暂停广告活动", "关闭广告活动"]) or "campaign pause" in lowered:
        return "pause_campaign"
    if any(token in text for token in ["暂停关键词", "暂停这个词", "暂停ASIN", "暂停投放", "关闭这个投放"]):
        return "pause_target"
    if any(token in text for token in ["首页首位", "商品页面", "其余位置", "广告位百分比", "修改广告位"]) or any(token in lowered for token in ["top of search", "product pages", "rest of search", "placement"]):
        return "update_placement"
    if any(token in text for token in ["预算改为", "预算改成", "预算加到", "预算降低到", "修改预算"]) or "daily budget" in lowered:
        return "update_campaign_budget"
    if any(token in text for token in ["修改ASIN定投竞价", "修改ASIN竞价", "修改商品投放竞价", "修改商品定投竞价", "ASIN定投竞价", "商品投放竞价"]) or "product targeting bid" in lowered or "asin bid" in lowered:
        return "update_product_targeting_bid"
    if any(token in text for token in ["竞价改为", "修改关键词竞价", "改关键词竞价", "关键词竞价改成", "这些词加价", "这些词降价", "修改竞价"]) or "bid改成" in text or "bid改为" in lowered:
        return "update_keyword_bid"
    return ""


def _parse_fields(body: str) -> dict[str, object]:
    fields: dict[str, object] = {}
    collecting = ""
    for line in body.split("\n"):
        stripped = line.strip()
        if not stripped:
            continue
        pair = re.match(r"^([^：:]+)[：:]\s*(.*)$", stripped)
        if pair:
            key = _field_key(pair.group(1))
            value = pair.group(2).strip()
            if key in {"keywords", "negative_keywords", "asins"}:
                collecting = key
                fields.setdefault(key, [])
                if value:
                    fields[key].extend(_split_items(value))
            elif key:
                collecting = ""
                if key == "match_type":
                    fields[key] = _canonical_match_type(value)
                elif key == "bidding_strategy":
                    fields[key] = _canonical_bidding_strategy(value)
                elif key == "state":
                    fields[key] = _canonical_state(value)
                elif key == "targeting_type":
                    fields[key] = _canonical_targeting_type(value)
                elif key == "placement":
                    fields[key] = _canonical_placement(value)
                elif key in {"budget", "bid", "percentage"}:
                    fields[key] = _clean_number_text(value)
                else:
                    fields[key] = _clean_text(value)
            else:
                collecting = ""
        elif collecting:
            fields.setdefault(collecting, [])
            fields[collecting].extend(_split_items(stripped))
    return fields


def _field_key(label: str) -> str:
    mapping = {
        "sku": "sku",
        "操作": "operation",
        "operation": "operation",
        "广告活动": "campaign_name",
        "广告活动名称": "campaign_name",
        "campaign": "campaign_name",
        "campaignname": "campaign_name",
        "广告组合": "portfolio_name",
        "广告组合名称": "portfolio_name",
        "portfolio": "portfolio_name",
        "portfolioname": "portfolio_name",
        "广告组合编号": "portfolio_id",
        "广告组合id": "portfolio_id",
        "portfolioid": "portfolio_id",
        "广告组": "ad_group_name",
        "广告组名称": "ad_group_name",
        "adgroup": "ad_group_name",
        "adgroupname": "ad_group_name",
        "匹配方式": "match_type",
        "matchtype": "match_type",
        "投放类型": "targeting_type",
        "竞价方式": "bidding_strategy",
        "竞价方案": "bidding_strategy",
        "竞价策略": "bidding_strategy",
        "biddingstrategy": "bidding_strategy",
        "状态": "state",
        "state": "state",
        "预算": "budget",
        "budget": "budget",
        "dailybudget": "budget",
        "竞价": "bid",
        "出价": "bid",
        "bid": "bid",
        "关键词": "keywords",
        "keyword": "keywords",
        "asin": "asins",
        "asinbid": "asins",
        "定投asin": "asins",
        "商品投放": "asins",
        "商品定投": "asins",
        "商品投放asin": "asins",
        "拓展商品投放": "asins",
        "扩展商品投放": "asins",
        "否定词": "negative_keywords",
        "否定关键词": "negative_keywords",
        "广告位": "placement",
        "广告位置": "placement",
        "placement": "placement",
        "百分比": "percentage",
        "广告位百分比": "percentage",
        "percentage": "percentage",
    }
    return mapping.get(_normalize(label), "")


def _split_items(value: str) -> list[str]:
    return [item for item in (_clean_text(raw) for raw in re.split(r"[,，;；\n]+", value)) if item]


def _detect_match_type(text: str) -> str:
    lowered = text.lower()
    if "精准" in text or "exact" in lowered:
        return "exact"
    if "广泛" in text or "broad" in lowered:
        return "broad"
    if "词组" in text or "phrase" in lowered:
        return "phrase"
    return "phrase"


def _extract_inline_keyword(text: str, match_type: str) -> str:
    match = re.search(r"(?:开个|新建|开)\s*(.+?)\s*(?:精准|词组|广泛|exact|phrase|broad|广告|关键词集合)", text, flags=re.IGNORECASE)
    candidate = _clean_text(match.group(1)) if match else ""
    return "" if _normalize(candidate) in {"关键词", "关键词广告", "广告"} else candidate


def _extract_inline_negative_keywords(text: str, action_type: str) -> list[str]:
    marker = "否定精准" if action_type == "add_negative_exact" else "否定词组"
    if marker not in text:
        return []
    before = text.split(marker, 1)[0]
    after = text.split(marker, 1)[-1]
    after = re.sub(r"^(：|:|为|关键词|否定词|添加)", "", after).strip()
    if after and not _contains_field_labels(after):
        return _split_items(after)
    before = re.sub(r"^(把|将|添加|关键词|否定词)", "", before.strip()).strip()
    return _split_items(before) if before and not _contains_field_labels(before) else []


def _extract_inline_campaign(text: str) -> str:
    for pattern in [r"广告活动[：:][ \t]*([^\n，。,]+)", r"分析[ \t]*([^\n，。,]+)"]:
        match = re.search(pattern, text)
        if match:
            return _clean_text(match.group(1))
    return ""


def _contains_field_labels(text: str) -> bool:
    return bool(re.search(r"(^|\n)\s*[^：:\n]{1,20}[：:]", text))


def _extract_sku(text: str) -> str:
    match = re.search(r"\bSKU\s*[：:= ]\s*([A-Za-z0-9._-]+)", text, flags=re.IGNORECASE)
    return match.group(1).strip() if match else ""


def _extract_number_after(text: str, label: str) -> str:
    match = re.search(rf"{re.escape(label)}\s*(?:改为|改成|加到|降低到|为|=|:|：)?\s*([$￥]?\s*[0-9,]+(?:\.[0-9]+)?\s*%?\s*(?:美金|美元)?)", text, flags=re.IGNORECASE)
    return _clean_number_text(match.group(1)) if match else ""


def _extract_placements(text: str) -> list[tuple[str, str]]:
    placements: list[tuple[str, str]] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        placement = _canonical_placement(line)
        if not placement:
            continue
        percentage = _clean_number_text(line)
        if percentage:
            placements.append((placement, percentage))
    fields = _parse_fields(text)
    placement = str(fields.get("placement", "") or "")
    percentage = str(fields.get("percentage", "") or "")
    if placement and percentage:
        placements.append((placement, percentage))
    unique: list[tuple[str, str]] = []
    seen = set()
    for placement, percentage in placements:
        key = (placement, percentage)
        if key not in seen:
            seen.add(key)
            unique.append(key)
    return unique


def _extract_pause_keywords(text: str) -> list[str]:
    keywords: list[str] = []
    collecting = False
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        key_match = re.match(r"^([^:：]+)[:：]\s*(.*)$", line)
        if key_match:
            key = _field_key(key_match.group(1))
            value = key_match.group(2).strip()
            collecting = key == "keywords"
            if collecting and value:
                keywords.extend(_split_items(value))
            continue
        if collecting:
            keywords.extend(_split_items(line))
    if keywords:
        return keywords
    for line in text.splitlines():
        stripped = line.strip()
        if stripped and not _field_key(stripped.split(":", 1)[0].split("：", 1)[0]) and not _line_starts_section(stripped):
            keywords.extend(_split_items(stripped))
    return [keyword for keyword in keywords if keyword and not _looks_like_asin(keyword)]


def _extract_asin(text: str) -> str:
    text = text.replace("“", '"').replace("”", '"')
    match = re.search(r'\basin\s*[：:=]\s*"?([A-Z0-9]{10})"?', text, flags=re.IGNORECASE)
    if match:
        return match.group(1).upper()
    match = re.search(r"\b(B0[A-Z0-9]{8})\b", text, flags=re.IGNORECASE)
    return match.group(1).upper() if match else ""


def _extract_asins(text: str) -> list[str]:
    normalized = text.replace("“", '"').replace("”", '"')
    values = re.findall(r'\basin\s*[：:=]\s*"?([A-Z0-9]{10})"?', normalized, flags=re.IGNORECASE)
    values.extend(re.findall(r"\b(B0[A-Z0-9]{8})\b", normalized, flags=re.IGNORECASE))
    unique: list[str] = []
    for value in values:
        asin = _canonical_asin(value)
        if asin and asin not in unique:
            unique.append(asin)
    return unique


def _extract_keyword_bid_pairs(text: str) -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        match = re.match(r"^(.+?)\s*(?:=|＝|->|→|改为|改成|竞价|出价|:|：)\s*([$￥]?\s*[0-9,]+(?:\.[0-9]+)?\s*(?:美金|美元)?)\s*$", line)
        if not match:
            continue
        keyword = _clean_text(match.group(1))
        bid = _clean_number_text(match.group(2))
        if not keyword or _field_key(keyword):
            continue
        pairs.append((keyword, bid))
    return pairs


def _extract_keyword_before_bid(text: str) -> str:
    if "竞价改为" in text:
        before = text.split("竞价改为", 1)[0]
        for line in reversed([part.strip() for part in before.splitlines() if part.strip()]):
            if re.search(r"(广告活动|广告组|SKU|预算|匹配方式)[：:]", line, flags=re.IGNORECASE):
                continue
            candidate = line
            candidate = re.sub(r"^(把|将|关键词)\s*", "", candidate).strip()
            candidate = re.sub(r"\s*(的)?$", "", candidate).strip()
            if candidate:
                return _clean_text(candidate)
    return ""


def _campaign_suffix(match_type: str) -> str:
    return {"exact": "EX", "phrase": "PHRASE", "broad": "BROAD"}.get(match_type, "PHRASE")


def _slug_keyword(keyword: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "-", keyword.lower()).strip("-")
    return slug[:60] or "keyword"


def _source_text(section: dict[str, str]) -> str:
    return f"{section['title']}\n{section['body']}".strip()


def _action_row(
    action_type: str,
    object_type: str,
    campaign_name: str,
    ad_group_name: str,
    sku: str,
    keyword: str,
    match_type: str,
    bidding_strategy: str,
    budget: str,
    bid: str,
    negative_type: str,
    state: str,
    reason: str,
    source_text: str,
    portfolio_name: str = "",
    portfolio_id: str = "",
) -> dict[str, object]:
    return {
        "action_id": "",
        "action_type": action_type,
        "object_type": object_type,
        "portfolio_id": _clean_text(portfolio_id),
        "portfolio_name": _clean_text(portfolio_name),
        "entity_id": "",
        "campaign_id": "",
        "ad_group_id": "",
        "ad_id": "",
        "keyword_id": "",
        "product_target_id": "",
        "campaign_name": _clean_text(campaign_name),
        "ad_group_name": _clean_text(ad_group_name),
        "sku": _clean_text(sku),
        "keyword": _clean_text(keyword),
        "match_type": _canonical_match_type(match_type),
        "bidding_strategy": _canonical_bidding_strategy(bidding_strategy),
        "budget": _clean_number_text(budget),
        "bid": _clean_number_text(bid),
        "negative_type": _canonical_negative_type(negative_type),
        "state": _canonical_state(state) or "enabled",
        "block_id": "",
        "block_index": "",
        "block_title": "",
        "reason": reason,
        "source_text": source_text,
        "source_row_index": "",
    }


def _assign_create_temp_ids(actions: pd.DataFrame) -> pd.DataFrame:
    actions = _clean_actions(actions).copy()
    campaign_ids: dict[str, str] = {}
    ad_group_ids: dict[tuple[str, str], str] = {}
    ad_ids: dict[tuple[str, str, str], str] = {}
    keyword_ids: dict[tuple[str, str, str, str, str], str] = {}
    campaign_count = 0
    ad_group_count = 0
    ad_count = 0
    keyword_count = 0

    for campaign_name in actions.loc[actions["action_type"].str.startswith("create_"), "campaign_name"].tolist():
        key = _norm_key(campaign_name)
        if key and key not in campaign_ids:
            campaign_count += 1
            campaign_ids[key] = f"tmp_campaign_{campaign_count:03d}"

    for row in actions[actions["action_type"].isin(["create_ad_group", "create_product_ad", "create_keyword_target", "create_product_target"])].to_dict("records"):
        campaign_key = _norm_key(row.get("campaign_name", ""))
        ad_group_key = _norm_key(row.get("ad_group_name", ""))
        key = (campaign_key, ad_group_key)
        if campaign_key and ad_group_key and key not in ad_group_ids:
            ad_group_count += 1
            ad_group_ids[key] = f"tmp_adgroup_{ad_group_count:03d}"

    for idx, row in actions.iterrows():
        action_type = str(row["action_type"])
        is_create = action_type.startswith("create_")
        is_negative_create = action_type in {"add_negative_exact", "add_negative_phrase"}
        if not is_create and not is_negative_create:
            continue
        campaign_key = _norm_key(row["campaign_name"])
        ad_group_key = _norm_key(row["ad_group_name"])
        if is_create and campaign_key:
            actions.at[idx, "campaign_id"] = campaign_ids.get(campaign_key, "")
        if action_type in {"create_ad_group", "create_product_ad", "create_keyword_target", "create_product_target"} and campaign_key and ad_group_key:
            actions.at[idx, "ad_group_id"] = ad_group_ids.get((campaign_key, ad_group_key), "")
        if action_type == "create_campaign":
            actions.at[idx, "entity_id"] = actions.at[idx, "campaign_id"]
        elif action_type == "create_ad_group":
            actions.at[idx, "entity_id"] = actions.at[idx, "ad_group_id"]
        elif action_type == "create_product_ad":
            ad_key = (campaign_key, ad_group_key, _norm_key(row["sku"]))
            if campaign_key and ad_group_key and ad_key not in ad_ids:
                ad_count += 1
                ad_ids[ad_key] = f"tmp_ad_{ad_count:03d}"
            actions.at[idx, "ad_id"] = ad_ids.get(ad_key, "")
            actions.at[idx, "entity_id"] = actions.at[idx, "ad_id"]
        elif action_type == "create_keyword_target":
            keyword_key = (campaign_key, ad_group_key, _norm_key(row["keyword"]), _norm_key(row["match_type"]), "keyword")
            if campaign_key and ad_group_key and keyword_key not in keyword_ids:
                keyword_count += 1
                keyword_ids[keyword_key] = f"tmp_keyword_{keyword_count:03d}"
            actions.at[idx, "keyword_id"] = keyword_ids.get(keyword_key, "")
            actions.at[idx, "entity_id"] = actions.at[idx, "keyword_id"]
        elif action_type == "create_product_target":
            target_key = (campaign_key, ad_group_key, _norm_key(row["keyword"]), "product_target")
            if campaign_key and ad_group_key and target_key not in keyword_ids:
                keyword_count += 1
                keyword_ids[target_key] = f"tmp_product_target_{keyword_count:03d}"
            actions.at[idx, "product_target_id"] = keyword_ids.get(target_key, "")
            actions.at[idx, "entity_id"] = actions.at[idx, "product_target_id"]
        elif is_negative_create:
            keyword_key = (campaign_key, ad_group_key, _norm_key(row["keyword"]), _norm_key(row["negative_type"]), "negative")
            if campaign_key and keyword_key not in keyword_ids:
                keyword_count += 1
                keyword_ids[keyword_key] = f"tmp_negative_keyword_{keyword_count:03d}"
            actions.at[idx, "keyword_id"] = keyword_ids.get(keyword_key, "")
            actions.at[idx, "entity_id"] = actions.at[idx, "keyword_id"]
    return actions


def _enrich_action(action: dict[str, object], indexes: BulkIndexes) -> dict[str, object]:
    enriched = dict(action)
    action_type = str(enriched.get("action_type", ""))
    if action_type.startswith("create_") and enriched.get("portfolio_name") and not enriched.get("portfolio_id"):
        enriched["portfolio_id"] = _portfolio_id_for(indexes, enriched.get("portfolio_name", ""))
    if action_type == "update_budget":
        enriched["campaign_id"] = _campaign_id_for(indexes, enriched.get("campaign_name", ""))
        enriched["entity_id"] = _campaign_entity_id_for(indexes, enriched.get("campaign_name", "")) or enriched["campaign_id"]
        enriched["source_row_index"] = _campaign_source_row_index_for(indexes, enriched.get("campaign_name", ""))
    elif action_type == "update_bid":
        enriched["campaign_id"] = _campaign_id_for(indexes, enriched.get("campaign_name", ""))
        if enriched.get("ad_group_name"):
            enriched["ad_group_id"] = _ad_group_id_for(indexes, enriched.get("campaign_name", ""), enriched.get("ad_group_name", ""))
        matches = _find_target(indexes, enriched["campaign_name"], enriched.get("ad_group_name", ""), enriched["keyword"])
        if len(matches) == 1:
            if not enriched.get("ad_group_name"):
                enriched["ad_group_name"] = str(matches.iloc[0]["ad_group_name"])
            if not enriched.get("match_type"):
                enriched["match_type"] = str(matches.iloc[0]["match_type"])
            enriched["campaign_id"] = str(matches.iloc[0].get("campaign_id", "") or "")
            enriched["ad_group_id"] = str(matches.iloc[0].get("ad_group_id", "") or "")
            enriched["keyword_id"] = str(matches.iloc[0].get("keyword_id", "") or "")
            enriched["product_target_id"] = str(matches.iloc[0].get("product_target_id", "") or "")
            enriched["entity_id"] = str(matches.iloc[0].get("entity_id", "") or "") or enriched["keyword_id"] or enriched["product_target_id"]
            enriched["source_row_index"] = str(matches.iloc[0].get("source_row_index", "") or "")
    elif action_type in PRODUCT_TARGETING_BID_ACTIONS:
        enriched["keyword"] = _canonical_asin(enriched.get("keyword", ""))
        matches = _find_product_target(indexes, enriched["campaign_name"], enriched.get("ad_group_name", ""), enriched["keyword"])
        if len(matches) == 1:
            if not enriched.get("campaign_name"):
                enriched["campaign_name"] = str(matches.iloc[0]["campaign_name"])
            if not enriched.get("ad_group_name"):
                enriched["ad_group_name"] = str(matches.iloc[0]["ad_group_name"])
            enriched["campaign_id"] = str(matches.iloc[0].get("campaign_id", "") or "")
            enriched["ad_group_id"] = str(matches.iloc[0].get("ad_group_id", "") or "")
            enriched["product_target_id"] = str(matches.iloc[0].get("product_target_id", "") or "")
            enriched["entity_id"] = str(matches.iloc[0].get("entity_id", "") or "") or enriched["product_target_id"]
            enriched["source_row_index"] = str(matches.iloc[0].get("source_row_index", "") or "")
    elif action_type == "update_placement":
        enriched["campaign_id"] = _campaign_id_for(indexes, enriched.get("campaign_name", ""))
        enriched["entity_id"] = enriched.get("campaign_id", "")
    elif action_type == "pause_campaign":
        enriched["campaign_id"] = _campaign_id_for(indexes, enriched.get("campaign_name", ""))
        enriched["entity_id"] = _campaign_entity_id_for(indexes, enriched.get("campaign_name", "")) or enriched["campaign_id"]
        enriched["source_row_index"] = _campaign_source_row_index_for(indexes, enriched.get("campaign_name", ""))
    elif action_type == "pause_target":
        target_value = str(enriched.get("keyword", ""))
        if _looks_like_asin(target_value):
            enriched["keyword"] = _canonical_asin(target_value)
            matches = _find_product_target(indexes, enriched["campaign_name"], enriched.get("ad_group_name", ""), enriched["keyword"])
            if len(matches) == 1:
                if not enriched.get("ad_group_name"):
                    enriched["ad_group_name"] = str(matches.iloc[0]["ad_group_name"])
                enriched["campaign_id"] = str(matches.iloc[0].get("campaign_id", "") or "")
                enriched["ad_group_id"] = str(matches.iloc[0].get("ad_group_id", "") or "")
                enriched["product_target_id"] = str(matches.iloc[0].get("product_target_id", "") or "")
                enriched["entity_id"] = str(matches.iloc[0].get("entity_id", "") or "") or enriched["product_target_id"]
                enriched["source_row_index"] = str(matches.iloc[0].get("source_row_index", "") or "")
        else:
            matches = _find_target(indexes, enriched["campaign_name"], enriched.get("ad_group_name", ""), enriched["keyword"])
            if len(matches) == 1:
                if not enriched.get("ad_group_name"):
                    enriched["ad_group_name"] = str(matches.iloc[0]["ad_group_name"])
                enriched["campaign_id"] = str(matches.iloc[0].get("campaign_id", "") or "")
                enriched["ad_group_id"] = str(matches.iloc[0].get("ad_group_id", "") or "")
                enriched["keyword_id"] = str(matches.iloc[0].get("keyword_id", "") or "")
                enriched["entity_id"] = str(matches.iloc[0].get("entity_id", "") or "") or enriched["keyword_id"]
                enriched["source_row_index"] = str(matches.iloc[0].get("source_row_index", "") or "")
    elif action_type in {"add_negative_exact", "add_negative_phrase"}:
        campaign_id = _campaign_id_for(indexes, enriched.get("campaign_name", ""))
        enriched["campaign_id"] = campaign_id
        if enriched.get("ad_group_name"):
            enriched["ad_group_id"] = _ad_group_id_for(indexes, enriched.get("campaign_name", ""), enriched.get("ad_group_name", ""))
        else:
            ad_group = _single_ad_group_for(indexes, enriched.get("campaign_name", ""))
            if ad_group:
                enriched["ad_group_name"] = ad_group.get("ad_group_name", "")
                enriched["ad_group_id"] = ad_group.get("ad_group_id", "")
        if not enriched.get("entity_id"):
            enriched["entity_id"] = enriched.get("keyword_id", "")
    return enriched


def _find_bulk_sheet(workbook) -> tuple[object | None, int | None, list[str]]:
    best = (None, None, [], 0)
    for sheet in workbook.worksheets:
        row_idx, headers, score = _find_header_row(sheet)
        if "商品推广" in sheet.title or "Sponsored Products" in sheet.title:
            score += 5
        if score > best[3]:
            best = (sheet, row_idx, headers, score)
    if best[3] < 4:
        return None, None, []
    return best[0], best[1], best[2]


def _find_header_row(sheet) -> tuple[int | None, list[str], int]:
    best_row = None
    best_headers: list[str] = []
    best_score = 0
    for row_idx in range(1, min(sheet.max_row, 30) + 1):
        headers = ["" if cell.value is None else str(cell.value).strip() for cell in sheet[row_idx]]
        normalized_headers = {_normalize(header) for header in headers if header}
        score = 0
        for aliases in BULK_HEADER_ALIASES.values():
            if any(_normalize(alias) in normalized_headers for alias in aliases):
                score += 1
        if score > best_score:
            best_row = row_idx
            best_headers = headers
            best_score = score
    return best_row, best_headers, best_score


def _build_header_map(headers: list[str]) -> dict[str, int]:
    normalized_headers = {_normalize(header): idx for idx, header in enumerate(headers) if header}
    header_map = dict(normalized_headers)
    for field, aliases in BULK_HEADER_ALIASES.items():
        for alias in aliases:
            idx = normalized_headers.get(_normalize(alias))
            if idx is not None:
                header_map[field] = idx
                break
    return header_map


def _sheet_to_dataframe(sheet, header_row_idx: int, headers: list[str]) -> pd.DataFrame:
    rows = []
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(value not in ("", None) for value in row):
            continue
        values = list(row[: len(headers)])
        values.extend([None] * (len(headers) - len(values)))
        rows.append(values[: len(headers)])
    return pd.DataFrame(rows, columns=headers)


def _capture_style_row(sheet, header_row_idx: int) -> list[object]:
    source_idx = header_row_idx + 1 if sheet.max_row > header_row_idx else header_row_idx
    return [copy.copy(cell._style) for cell in sheet[source_idx]]


def _extract_labels(workbook, data: pd.DataFrame, header_map: dict[str, int]) -> dict[str, dict[str, str]]:
    labels = _fallback_labels()
    _extract_config_labels(workbook, labels)
    for _, row in data.iterrows():
        for group, field, canonical_func in [
            ("entity", "entity", _canonical_entity),
            ("operation", "operation", _canonical_operation),
            ("state", "state", _canonical_state),
            ("match_type", "match_type", _canonical_match_type),
            ("targeting_type", "targeting_type", _canonical_targeting_type),
            ("bidding_strategy", "bidding_strategy", _canonical_bidding_strategy),
        ]:
            raw = str(_row_value(row, header_map, field)).strip()
            canonical = canonical_func(raw)
            if group == "entity" and canonical == "keyword_target" and (
                _row_value(row, header_map, "product_target_id") or _looks_like_asin(_target_value(row, header_map))
            ):
                canonical = "product_target"
            if raw and canonical and not _has_cjk(raw):
                labels[group][canonical] = raw
        product = str(_row_value(row, header_map, "product")).strip()
        if product and not _has_cjk(product):
            labels["product"]["sponsored_products"] = product
    return labels


def _extract_config_labels(workbook, labels: dict[str, dict[str, str]]) -> None:
    config_groups = {
        "sponsoredproductsproductnames": ("product", _canonical_product),
        "sponsoredproductsentitynames": ("entity", _canonical_entity),
        "sponsoredproductsoperationnames": ("operation", _canonical_operation),
        "sponsoredproductscreatecampaigntargetingtypes": ("targeting_type", _canonical_targeting_type),
        "sponsoredproductscreatecampaignstates": ("state", _canonical_state),
        "sponsoredproductscreatecampaignstrategys": ("bidding_strategy", _canonical_bidding_strategy),
        "sponsoredproductsupdatecampaignstrategys": ("bidding_strategy", _canonical_bidding_strategy),
        "sponsoredproductscreatekeywordmatchtypes": ("match_type", _canonical_match_type),
        "sponsoredproductscreatenegativekeywordmatchtypes": ("match_type", _canonical_match_type),
        "sponsoredproductscreatecampaignnegativekeywordmatchtypes": ("match_type", _canonical_match_type),
    }
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            config_key = _normalize(row[0])
            group_info = config_groups.get(config_key)
            if not group_info:
                continue
            group, canonical_func = group_info
            for raw in row[1:]:
                raw_text = str(raw or "").strip()
                if not raw_text or _has_cjk(raw_text):
                    continue
                canonical = canonical_func(raw_text)
                if canonical:
                    labels[group][canonical] = raw_text


def _fallback_labels() -> dict[str, dict[str, str]]:
    return {
        "product": {"sponsored_products": "Sponsored Products"},
        "operation": {"create": "Create", "update": "Update", "archive": "Archive"},
        "state": {"enabled": "enabled", "paused": "paused", "archived": "archived"},
        "targeting_type": {"manual": "Manual", "auto": "Auto"},
        "bidding_strategy": {
            "fixed": "Fixed bids",
            "dynamic_up_down": "Dynamic bids - up and down",
            "dynamic_down_only": "Dynamic bids - down only",
            "up_down": "Dynamic bids - up and down",
            "down_only": "Dynamic bids - down only",
        },
        "entity": {"campaign": "Campaign", "ad_group": "Ad Group", "product_ad": "Product Ad", "keyword_target": "Keyword", "product_target": "Product Targeting", "placement": "Bidding Adjustment", "negative_keyword": "Negative Keyword", "campaign_negative_keyword": "Campaign Negative Keyword"},
        "match_type": {"exact": "exact", "phrase": "phrase", "broad": "broad", "negative exact": "negativeExact", "negative phrase": "negativePhrase"},
    }


def _validate_template(template: BulkTemplate) -> list[dict[str, object]]:
    missing = [field for field in BASE_REQUIRED_BULK_FIELDS if field not in template.header_map]
    if missing:
        return [_validation_row("Bulk 列校验", "error", f"模板缺少基础列：{', '.join(missing)}")]
    return [_validation_row("Bulk 列校验", "ok", "已识别基础列")]


def _clear_data_rows(sheet, header_row_idx: int) -> None:
    if sheet.max_row > header_row_idx:
        sheet.delete_rows(header_row_idx + 1, sheet.max_row - header_row_idx)


def _apply_style_to_row(sheet, row_idx: int, styles: list[object]) -> None:
    for col_idx, style in enumerate(styles, start=1):
        if col_idx <= sheet.max_column:
            sheet.cell(row=row_idx, column=col_idx)._style = copy.copy(style)


def _set_row_value(row_values: list[object], header_map: dict[str, int], field: str, value: object) -> None:
    idx = header_map.get(field)
    if idx is None or idx >= len(row_values) or value in ("", None):
        return
    row_values[idx] = _coerce_value(value)


def _coerce_value(value: object) -> object:
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if re.fullmatch(r"-?\d+(\.\d+)?", text):
        try:
            return float(text)
        except ValueError:
            return text
    return text


def _label(template: BulkTemplate, group: str, key: object) -> str:
    return template.labels.get(group, {}).get(str(key), str(key))


def _header_name(template: BulkTemplate, field: str) -> str:
    idx = template.header_map.get(field)
    if idx is not None and idx < len(template.headers):
        return template.headers[idx]
    return field


def _campaign_exists(indexes: BulkIndexes, campaign_name: str) -> bool:
    if indexes.campaigns.empty:
        return False
    return indexes.campaigns["campaign_name"].astype(str).map(_norm_key).eq(_norm_key(campaign_name)).any()


def _campaign_id_for(indexes: BulkIndexes, campaign_name: object) -> str:
    if indexes.campaigns.empty:
        return ""
    mask = indexes.campaigns["campaign_name"].astype(str).map(_norm_key).eq(_norm_key(campaign_name))
    values = indexes.campaigns.loc[mask, "campaign_id"].astype(str).str.strip()
    values = values[values != ""]
    return values.iloc[0] if not values.empty else ""


def _portfolio_id_for(indexes: BulkIndexes, portfolio_name: object) -> str:
    if indexes.portfolios.empty:
        return ""
    df = indexes.portfolios
    name_mask = df["portfolio_name"].astype(str).map(_norm_key).eq(_norm_key(portfolio_name))
    values = df.loc[name_mask, "portfolio_id"].astype(str).str.strip()
    values = values[values != ""]
    if not values.empty:
        return values.iloc[0]
    id_mask = df["portfolio_id"].astype(str).map(_norm_key).eq(_norm_key(portfolio_name))
    values = df.loc[id_mask, "portfolio_id"].astype(str).str.strip()
    values = values[values != ""]
    return values.iloc[0] if not values.empty else ""


def _campaign_entity_id_for(indexes: BulkIndexes, campaign_name: object) -> str:
    if indexes.campaigns.empty:
        return ""
    mask = indexes.campaigns["campaign_name"].astype(str).map(_norm_key).eq(_norm_key(campaign_name))
    values = indexes.campaigns.loc[mask, "entity_id"].astype(str).str.strip()
    values = values[values != ""]
    return values.iloc[0] if not values.empty else ""


def _campaign_source_row_index_for(indexes: BulkIndexes, campaign_name: object) -> str:
    if indexes.campaigns.empty or "source_row_index" not in indexes.campaigns.columns:
        return ""
    df = indexes.campaigns
    mask = df["campaign_name"].astype(str).map(_norm_key).eq(_norm_key(campaign_name))
    if "entity" in df.columns:
        campaign_rows = df[mask & df["entity"].astype(str).eq("campaign")]
        values = campaign_rows["source_row_index"].astype(str).str.strip()
        values = values[values != ""]
        if not values.empty:
            return values.iloc[0]
    values = df.loc[mask, "source_row_index"].astype(str).str.strip()
    values = values[values != ""]
    return values.iloc[0] if not values.empty else ""


def _ad_group_exists(indexes: BulkIndexes, campaign_name: str, ad_group_name: str) -> bool:
    if indexes.ad_groups.empty:
        return False
    df = indexes.ad_groups
    return (
        df["campaign_name"].astype(str).map(_norm_key).eq(_norm_key(campaign_name))
        & df["ad_group_name"].astype(str).map(_norm_key).eq(_norm_key(ad_group_name))
    ).any()


def _ad_group_id_for(indexes: BulkIndexes, campaign_name: object, ad_group_name: object) -> str:
    if indexes.ad_groups.empty:
        return ""
    df = indexes.ad_groups
    mask = (
        df["campaign_name"].astype(str).map(_norm_key).eq(_norm_key(campaign_name))
        & df["ad_group_name"].astype(str).map(_norm_key).eq(_norm_key(ad_group_name))
    )
    values = df.loc[mask, "ad_group_id"].astype(str).str.strip()
    values = values[values != ""]
    return values.iloc[0] if not values.empty else ""


def _single_ad_group_for(indexes: BulkIndexes, campaign_name: object) -> dict[str, str]:
    if indexes.ad_groups.empty:
        return {}
    df = indexes.ad_groups[indexes.ad_groups["campaign_name"].astype(str).map(_norm_key).eq(_norm_key(campaign_name))]
    if df.empty:
        return {}
    unique = df[["ad_group_name", "ad_group_id"]].drop_duplicates()
    names = [name for name in unique["ad_group_name"].astype(str).str.strip().unique().tolist() if name]
    if len(names) != 1:
        return {}
    row = unique[unique["ad_group_name"].astype(str).str.strip().eq(names[0])].iloc[0]
    return {"ad_group_name": str(row.get("ad_group_name", "") or ""), "ad_group_id": str(row.get("ad_group_id", "") or "")}


def _find_target(indexes: BulkIndexes, campaign_name: str, ad_group_name: str, keyword: str) -> pd.DataFrame:
    if indexes.targets.empty:
        return indexes.targets
    df = indexes.targets
    if "entity" in df.columns:
        df = df[df["entity"].astype(str).ne("product_target")]
    if "product_target_id" in df.columns:
        df = df[df["product_target_id"].astype(str).str.strip().eq("")]
    asin_mask = df["keyword"].astype(str).map(_looks_like_asin).astype(bool)
    df = df.loc[~asin_mask]
    if df.empty:
        return df
    mask = df["campaign_name"].astype(str).map(_norm_key).eq(_norm_key(campaign_name)) & df["keyword"].astype(str).map(_norm_key).eq(_norm_key(keyword))
    if ad_group_name:
        mask &= df["ad_group_name"].astype(str).map(_norm_key).eq(_norm_key(ad_group_name))
    return df[mask]


def _find_product_target(indexes: BulkIndexes, campaign_name: str, ad_group_name: str, asin: str) -> pd.DataFrame:
    product_rows = _product_target_rows(indexes)
    if product_rows.empty:
        return product_rows
    asin_key = _canonical_asin(asin)
    asin_mask = product_rows["keyword"].astype(str).apply(lambda value: _target_contains_asin(value, asin=asin_key))
    rows = product_rows[asin_mask]
    if rows.empty:
        return rows

    campaign_mask = rows["campaign_name"].astype(str).map(_norm_key).eq(_norm_key(campaign_name))
    ad_group_mask = rows["ad_group_name"].astype(str).map(_norm_key).eq(_norm_key(ad_group_name))
    if campaign_name and ad_group_name:
        exact = rows[campaign_mask & ad_group_mask]
        if not exact.empty:
            return exact
    if campaign_name:
        campaign_matches = rows[campaign_mask]
        if not campaign_matches.empty:
            return campaign_matches
    return rows


def _product_target_rows(indexes: BulkIndexes) -> pd.DataFrame:
    if indexes.targets.empty:
        return indexes.targets
    df = indexes.targets
    mask = pd.Series(False, index=df.index)
    if "entity" in df.columns:
        mask |= df["entity"].astype(str).eq("product_target")
    if "product_target_id" in df.columns:
        mask |= df["product_target_id"].astype(str).str.strip().ne("")
    mask |= df["keyword"].astype(str).map(_looks_like_asin)
    return df[mask]


def _target_contains_asin(value: object, *, asin: str) -> bool:
    if not asin:
        return False
    text = str(value or "").upper()
    return asin in text or _canonical_asin(value) == asin


def _product_target_not_found_reason(indexes: BulkIndexes) -> str:
    if _product_target_rows(indexes).empty:
        return "上传的 Bulk 表缺少 Product Targeting 行，请重新下载包含 Product Targeting 的 Bulk 表。"
    return "该 ASIN 定投未在上传的 Bulk 表中找到；可能是原 Bulk 表不包含该 ASIN 定投，或该 ASIN 定投尚未创建。建议重新下载包含 Product Targeting 的 Bulk 表，或者改用“新建ASIN定投”。"


def _negative_exists(indexes: BulkIndexes, campaign_name: str, ad_group_name: str, keyword: str, negative_type: str) -> bool:
    if indexes.negatives.empty:
        return False
    df = indexes.negatives
    mask = (
        df["campaign_name"].astype(str).map(_norm_key).eq(_norm_key(campaign_name))
        & df["keyword"].astype(str).map(_norm_key).eq(_norm_key(keyword))
        & df["negative_type"].astype(str).eq(negative_type)
    )
    if ad_group_name:
        mask &= df["ad_group_name"].astype(str).map(_norm_key).eq(_norm_key(ad_group_name))
    return mask.any()


def _target_value(row: pd.Series, header_map: dict[str, int]) -> str:
    keyword = _row_value(row, header_map, "keyword_text")
    expression = _row_value(row, header_map, "product_targeting_expression")
    return _clean_target_text(keyword or expression)


def _entity_id_for_row(row: pd.Series, header_map: dict[str, int]) -> str:
    entity = _canonical_entity(_row_value(row, header_map, "entity"))
    explicit = _row_value(row, header_map, "entity_id")
    if explicit:
        return str(explicit).strip()
    if entity == "campaign":
        return str(_row_value(row, header_map, "campaign_id")).strip()
    if entity == "ad_group":
        return str(_row_value(row, header_map, "ad_group_id")).strip()
    if entity == "product_ad":
        return str(_row_value(row, header_map, "ad_id")).strip()
    if entity in {"keyword_target", "negative_keyword", "campaign_negative_keyword"}:
        return str(_row_value(row, header_map, "keyword_id") or _row_value(row, header_map, "product_target_id")).strip()
    if entity == "product_target":
        return str(_row_value(row, header_map, "product_target_id")).strip()
    return ""


def _canonical_product(value: object) -> str:
    normalized = _normalize(value)
    if normalized in {"sponsoredproducts", "商品推广"}:
        return "sponsored_products"
    return ""


def _canonical_entity(value: object) -> str:
    normalized = _normalize(value)
    if normalized in {"campaign", "广告活动"}:
        return "campaign"
    if normalized in {"adgroup", "广告组"}:
        return "ad_group"
    if normalized in {"productad", "商品广告"}:
        return "product_ad"
    if normalized in {"keyword", "keywords", "keywordtarget", "keywordtargeting", "biddablekeyword", "sponsoredproductskeyword", "关键词", "关键词投放", "关键词定位", "关键词定向", "关键词目标"}:
        return "keyword_target"
    if normalized in {"targeting", "target", "投放", "定位", "定向"}:
        return "keyword_target"
    if normalized in {"producttargeting", "producttarget", "商品投放", "商品定位", "商品定向"}:
        return "product_target"
    if normalized in {"negativekeyword", "否定关键词"}:
        return "negative_keyword"
    if normalized in {"campaignnegativekeyword", "广告活动否定关键词"}:
        return "campaign_negative_keyword"
    return ""


def _canonical_operation(value: object) -> str:
    normalized = _normalize(value)
    if normalized in {"create", "创建"}:
        return "create"
    if normalized in {"update", "更新"}:
        return "update"
    if normalized in {"archive", "archived", "归档"}:
        return "archive"
    return ""


def _canonical_state(value: object) -> str:
    normalized = _normalize(value)
    if normalized in {"enabled", "enable", "已启用", "启用", "开启", "开"}:
        return "enabled"
    if normalized in {"paused", "pause", "已暂停", "暂停", "关闭", "关"}:
        return "paused"
    if normalized in {"archived", "archive", "已归档", "归档"}:
        return "archived"
    return str(value or "").strip().lower()


def _canonical_targeting_type(value: object) -> str:
    normalized = _normalize(value)
    if normalized in {"manual", "manualtargeting", "手动", "手动投放"}:
        return "manual"
    if normalized in {"auto", "automatic", "automatictargeting", "自动", "自动投放"}:
        return "auto"
    return ""


def _canonical_placement(value: object) -> str:
    normalized = _normalize(value)
    if normalized in {"topofsearch", "topsearch", "首页首位", "搜索结果顶部", "搜索顶部"} or "topofsearch" in normalized or "首页首位" in normalized:
        return "top_of_search"
    if normalized in {"productpages", "productpage", "商品页面", "商品页"} or "productpages" in normalized or "商品页面" in normalized:
        return "product_pages"
    if normalized in {"restofsearch", "otherplacements", "其余位置", "搜索结果其余位置", "其他位置"} or "restofsearch" in normalized or "其余位置" in normalized:
        return "rest_of_search"
    return ""


def _placement_display(value: object) -> str:
    return {
        "top_of_search": "Top of Search",
        "product_pages": "Product Pages",
        "rest_of_search": "Rest of Search",
    }.get(str(value or ""), str(value or ""))


def _placement_label(template: BulkTemplate, value: object) -> str:
    return _placement_display(_canonical_placement(value) or value)


def _canonical_bidding_strategy(value: object) -> str:
    normalized = _normalize(value)
    if normalized in {"fixed", "fixedbid", "fixedbids", "固定", "固定竞价"}:
        return "fixed"
    if normalized in {
        "dynamicbidsupanddown",
        "upanddown",
        "updown",
        "升高或者降低",
        "升高或降低",
        "提高和降低",
        "提高或降低",
        "动态竞价升高或者降低",
        "动态竞价提高和降低",
    }:
        return "dynamic_up_down"
    if normalized in {
        "dynamicbidsdownonly",
        "downonly",
        "down",
        "仅降低",
        "只降低",
        "动态竞价仅降低",
        "动态竞价只降低",
    }:
        return "dynamic_down_only"
    return ""


def _canonical_match_type(value: object) -> str:
    normalized = _normalize(value)
    if normalized in {"exact", "精准", "精准匹配"}:
        return "exact"
    if normalized in {"phrase", "词组", "词组匹配"}:
        return "phrase"
    if normalized in {"broad", "广泛", "广泛匹配"}:
        return "broad"
    if normalized in {"negativeexact", "negativeexactmatch", "否定精准", "否定精准匹配"}:
        return "negative exact"
    if normalized in {"negativephrase", "negativephrasematch", "否定词组", "否定词组匹配"}:
        return "negative phrase"
    return str(value or "").strip().lower()


def _canonical_negative_type(value: object) -> str:
    canonical = _canonical_match_type(value)
    return canonical if canonical in {"negative exact", "negative phrase"} else ""


def _read_bytes(file: BinaryIO | bytes | str | Path) -> bytes:
    if isinstance(file, bytes):
        return file
    if isinstance(file, (str, Path)):
        return Path(file).read_bytes()
    if hasattr(file, "getvalue"):
        return file.getvalue()
    return file.read()


def _row_value(row: pd.Series, header_map: dict[str, int], field: str) -> object:
    idx = header_map.get(field)
    if idx is None or idx >= len(row):
        return ""
    value = row.iloc[idx]
    return "" if pd.isna(value) else value


def _clean_actions(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return _empty_actions()
    cleaned = df.copy()
    for column in ACTION_COLUMNS:
        if column not in cleaned.columns:
            cleaned[column] = ""
    cleaned = cleaned[ACTION_COLUMNS].fillna("")
    for column in ACTION_COLUMNS:
        cleaned[column] = cleaned[column].astype(str).str.strip()
    return cleaned


def _clean_skipped(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return _empty_skipped()
    for column in SKIPPED_COLUMNS:
        if column not in df.columns:
            df[column] = ""
    return df[SKIPPED_COLUMNS].fillna("").astype(str)


def _assign_action_ids(df: pd.DataFrame) -> pd.DataFrame:
    df = _clean_actions(df)
    for idx in range(len(df)):
        if not df.at[idx, "action_id"]:
            df.at[idx, "action_id"] = f"A{idx + 1:03d}"
    return df


def _empty_actions() -> pd.DataFrame:
    return pd.DataFrame(columns=ACTION_COLUMNS)


def _empty_skipped() -> pd.DataFrame:
    return pd.DataFrame(columns=SKIPPED_COLUMNS)


def _empty_indexes() -> BulkIndexes:
    return BulkIndexes(
        campaigns=pd.DataFrame(columns=["entity", "campaign_name", "entity_id", "campaign_id", "portfolio_id", "portfolio_name", "source_row_index"]),
        ad_groups=pd.DataFrame(columns=["campaign_name", "campaign_id", "ad_group_name", "entity_id", "ad_group_id"]),
        targets=pd.DataFrame(columns=["entity", "campaign_name", "campaign_id", "ad_group_name", "ad_group_id", "entity_id", "keyword_id", "product_target_id", "keyword", "match_type", "source_row_index"]),
        negatives=pd.DataFrame(columns=["campaign_name", "campaign_id", "ad_group_name", "ad_group_id", "entity_id", "keyword_id", "keyword", "negative_type"]),
        portfolios=pd.DataFrame(columns=["portfolio_name", "portfolio_id"]),
    )


def _validation_row(check: str, status: str, detail: str, action_id: str = "") -> dict[str, object]:
    return {
        "check": check,
        "status": status,
        "detail": detail,
        "action_id": action_id,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }


def _failure_reasons(rows: list[dict[str, object]]) -> list[str]:
    reasons = []
    for row in rows:
        if row.get("status") == "error":
            detail = str(row.get("detail", ""))
            if detail and detail not in reasons:
                reasons.append(detail)
    return reasons


def _has_errors(rows: list[dict[str, object]]) -> bool:
    return any(row.get("status") == "error" for row in rows)


def _dedupe_skipped(records: list[dict[str, object]]) -> list[dict[str, object]]:
    seen = set()
    deduped = []
    for record in records:
        key = (record.get("action_id"), record.get("skip_reason"))
        if key not in seen:
            seen.add(key)
            deduped.append(record)
    return deduped


def _joined_unique(values: list[str]) -> str:
    unique = []
    for value in values:
        if value and value not in unique:
            unique.append(value)
    return "；".join(unique)


def _clean_target_text(value: object) -> str:
    text = str(value or "").replace("“", '"').replace("”", '"').strip()
    keyword_match = re.search(r'\b(?:keyword|keywordtext|keyword_text)\s*=\s*"([^"]+)"', text, flags=re.IGNORECASE)
    if keyword_match:
        return keyword_match.group(1).strip()
    keyword_match = re.search(r"\b(?:keyword|keywordtext|keyword_text)\s*=\s*([^,;\]\n]+)", text, flags=re.IGNORECASE)
    if keyword_match:
        return _clean_text(keyword_match.group(1))
    match = re.search(r'asin\s*[：:=]\s*"?([A-Z0-9]{10})"?', text, flags=re.IGNORECASE)
    return match.group(1).upper() if match else text


def _canonical_asin(value: object) -> str:
    text = _clean_target_text(value).upper()
    match = re.search(r"\b(B0[A-Z0-9]{8})\b", text)
    return match.group(1) if match else _clean_text(text)


def _looks_like_asin(value: object) -> bool:
    return bool(re.fullmatch(r"B0[A-Z0-9]{8}", _canonical_asin(value)))


def _is_empty_value(value: object) -> bool:
    normalized = _normalize(value)
    return normalized in {"", "无", "不改", "空", "没有", "none", "null", "na", "n/a", "no"}


def _clean_text(value: object) -> str:
    text = str(value or "").strip(" ：:，,。.\n\t")
    return "" if _is_empty_value(text) else text


def _clean_number_text(value: object) -> str:
    if _is_empty_value(value):
        return ""
    text = str(value or "").replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return match.group(0) if match else ""


def _num(value: object) -> float:
    try:
        text = _clean_number_text(value)
        return float(text) if text else 0.0
    except ValueError:
        return 0.0


def _norm_key(value: object) -> str:
    return _normalize(str(value or ""))


def _normalize(value: object) -> str:
    normalized = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", normalized)


def _has_cjk(value: str) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", value))
